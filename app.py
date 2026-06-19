import os
import re
import unicodedata
from io import BytesIO
from difflib import get_close_matches

import streamlit as st

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential


# ============================================================
# 1) EXCEL TEMPLATE (15 COLUMNS) + STYLE
# ============================================================
COLETA_COLUMNS = [
    "Chassi Série",
    "Tag Frota",
    "Categoria",
    "Ponto de Coleta / Compartimento",
    "Horímetro/Km/Período",
    "Número do Frasco",
    "Data da Coleta",
    "Óleo trocado",
    "Volume adicionado",
    "Fabricante (Óleo)",
    "Modelo (Óleo)",
    "Fabricante/Modelo",
    "Viscosidade (Óleo)",
    "Horas/Km do Fluído",
    "Comentário",
    "Código externo",
]

# Internal-only fields fed straight from Azure: they are combined into the visible
# "Fabricante/Modelo" column and then blanked in the Excel. Viscosidade is NO LONGER
# here — it is now its own visible column (split from the old "Descrição do Óleo").
FORCE_EMPTY_IN_EXCEL = {"Fabricante (Óleo)", "Modelo (Óleo)"}

# Below this confidence (%), a cell is highlighted orange in the Excel so the
# person knows to double-check it. Default applies to all columns; the overrides
# use a LOWER bar for fields where a modest confidence is normal and trustworthy:
#  - Data da Coleta: handwritten dates always read with low Azure confidence even
#    when correct; the format rule already validates them, so don't flag valid ones.
#  - Óleo trocado: a Sim/Não detected from pen-ink is reliable at ~75-88%; flagging
#    it at the 90% bar would mark almost every row as "review" (noise).
CONFIDENCE_REVIEW_THRESHOLD = 90.0
CONFIDENCE_REVIEW_OVERRIDES = {
    "Data da Coleta": 60.0,
    "Óleo trocado":   70.0,
}


def _review_threshold(col_name: str) -> float:
    return CONFIDENCE_REVIEW_OVERRIDES.get(col_name, CONFIDENCE_REVIEW_THRESHOLD)

COMPARTMENT_LABELS = [
    "Móvel",
    "Motor", "Comando Final Dir.", "Comando Final Esq.", "Comando de Giro",
    "Compressor", "Diesel", "Direção", "Eixo Dianteiro", "Eixo Traseiro",
    "Graxa", "Radiador", "Transmissão", "Hidráulico",
    "Industrial",
    "Bomba", "Sistema de Lubr.", "Redutor", "Sistema Térmico", "Turbina",
    "Ponto de Coleta", "Mancal", "Isolante",
    "Compressor Industrial", "Graxa Industrial", "Hidráulico Industrial",
    "Engrenagens", "Redução final Esq.", "Redução final Dir.", "Tanque",
]

# When one of these "category" labels is marked alongside a specific compartment,
# the category label is dropped in favour of the specific one.
CATEGORY_LABELS = {"Móvel", "Industrial"}
_PM_DEBUG = False   # set True to print per-mark debug info in _find_compartment_via_page_marks


REDIRECT_TO_DESCRICAO = set()  # all oil fields now consolidated in fields_to_record

COL_WIDTHS = {
    "Chassi Série": 22,
    "Tag Frota": 22,
    "Ponto de Coleta / Compartimento": 32,
    "Horímetro/Km/Período": 25,
    "Número do Frasco": 25,
    "Data da Coleta": 16,
    "Óleo trocado": 18,
    "Volume adicionado": 22,
    "Fabricante (Óleo)": 20,
    "Modelo (Óleo)": 18,
    "Fabricante/Modelo": 32,
    "Viscosidade (Óleo)": 18,
    "Horas/Km do Fluído": 22,
    "Comentário": 22,
    "Código externo": 18,
}

GREEN_FILL  = PatternFill("solid", fgColor="D9EAD3")
PINK_FILL   = PatternFill("solid", fgColor="F4CCCC")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
REVIEW_FILL = PatternFill("solid", fgColor="FFD699")   # low-confidence cell highlight

GREEN_COLS = {
    "Horímetro/Km/Período",
    "Data da Coleta",
    "Óleo trocado",
    "Volume adicionado",
    "Fabricante/Modelo",
    "Horas/Km do Fluído",
    "Comentário",
    "Código externo",
}
PINK_COLS = {
    "Número do Frasco",
    "Viscosidade (Óleo)",
}

DV_PONTO_COLETA = ["MOTOR", "REDUTOR", "TRANSMISSÃO", "DIFERENCIAL", "HIDRÁULICO", "COMPRESSOR", "RADIADOR", "OUTROS"]
DV_OLEO_TROCADO = ["Sim", "Não"]
DV_DESCRICAO    = ["SINTÉTICO", "MINERAL"]

CENTER_COLS = {
    "Número do Frasco",
    "Código externo",
    "Data da Coleta",
    "Óleo trocado",
    "Horímetro/Km/Período",
}


def build_excel_bytes(records: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coleta"
    wb.create_sheet("Referências")

    header_font  = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(COLETA_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 18)
    ws.row_dimensions[1].height = 22

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLETA_COLUMNS))}1"

    for r_idx, rec in enumerate(records, start=2):
        for c_idx, col_name in enumerate(COLETA_COLUMNS, start=1):
            value = "" if col_name in FORCE_EMPTY_IN_EXCEL else rec.get(col_name, "")
            ws.cell(row=r_idx, column=c_idx, value=value)

    max_row = max(2, ws.max_row)
    max_col = len(COLETA_COLUMNS)

    for c_idx, col_name in enumerate(COLETA_COLUMNS, start=1):
        fill = GREEN_FILL if col_name in GREEN_COLS else PINK_FILL if col_name in PINK_COLS else WHITE_FILL
        for r_idx in range(1, max_row + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill = fill
            if r_idx >= 2 and col_name in CENTER_COLS:
                cell.alignment = center_align

    # Realce de confiabilidade: pinta de laranja cada célula PREENCHIDA cujo score
    # composto (Azure + regra de plausibilidade + catálogo) ficou abaixo do limite
    # de revisão — sobrepõe a cor base da coluna. Células vazias ficam como estão.
    for r_idx, rec in enumerate(records, start=2):
        conf = rec.get("__confidence__", {}) or {}
        for c_idx, col_name in enumerate(COLETA_COLUMNS, start=1):
            if (
                col_name not in FORCE_EMPTY_IN_EXCEL
                and rec.get(col_name, "")
                and col_name in conf
                and conf[col_name] < _review_threshold(col_name)
            ):
                ws.cell(row=r_idx, column=c_idx).fill = REVIEW_FILL

    thin  = Side(style="thin",   color="B7B7B7")
    thick = Side(style="medium", color="808080")

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        for r in range(1, max_row + 1):
            cell = ws[f"{col_letter}{r}"]
            cell.border = Border(
                left=cell.border.left,
                right=thick,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    last_row = max(500, max_row)

    def add_list_validation(col_name: str, options: list[str]):
        if col_name not in COLETA_COLUMNS:
            return
        col_idx    = COLETA_COLUMNS.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True, showDropDown=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")

    add_list_validation("Ponto de Coleta / Compartimento", DV_PONTO_COLETA)
    add_list_validation("Óleo trocado",                    DV_OLEO_TROCADO)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ============================================================
# 2) NORMALIZATION + MAPPING (OCR KEYS -> EXCEL COLUMNS)
# ============================================================
def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9 /?]", "", s)
    return s.strip()


COL_NORM_MAP      = {_norm(c): c for c in COLETA_COLUMNS}
COL_NORMS         = list(COL_NORM_MAP.keys())
COMPARTMENT_OPTIONS = {_norm(label): label for label in COMPARTMENT_LABELS}

# ── EASYOCR FALLBACK ──────────────────────────────────────────────────────────
_easyocr_reader = None

def _get_easyocr_reader():
    """Lazy-load EasyOCR reader (cached). Returns None if unavailable."""
    global _easyocr_reader
    if _easyocr_reader is None:
        try:
            import easyocr
            _easyocr_reader = easyocr.Reader(["pt", "en"], gpu=False, verbose=False)
        except Exception:
            pass
    return _easyocr_reader


def _read_field_from_words(page, label_triggers: list[str],
                            search_below: bool = False) -> str:
    """
    Stage 1: find a field value from Azure page.words — text to the right
    of the label on the same line, stopping at the next form label.
    When search_below=True and nothing is found horizontally, also searches
    for the first line of text directly below the label (up to 0.55" down).
    """
    words = getattr(page, "words", None) or []
    if not words:
        return ""

    FORM_KEYWORDS = {
        "chassi", "serie", "serial", "tag", "frota",
        "data", "coleta", "horimetro", "horometro",
        "numero", "frasco", "oleo", "trocado",
        "volume", "adicionado", "fabricante", "viscosidade",
        "modelo", "descricao", "horas", "km", "fluido",
        "comentario", "codigo", "externo", "ponto",
        "compartimento", "periodo", "equipamento",
        "motor", "transmissao", "hidraulico", "comando", "giro",
        "compressor", "diesel", "direcao", "eixo", "dianteiro",
        "traseiro", "graca", "graxa", "redutor", "turbina",
        "bomba", "mancal", "isolante", "tanque", "radiador",
        "sistema", "lubr", "termico",
        # "Vol." / "adic." label abbreviations — prevents below-search picking up
        # the "Vol. fluido adic." row as a value for the Horas/km field
        "vol", "adic",
        # Short connectors: treated as label glue so Pass 3 can crawl through
        # "HORAS/KM DO FLUIDO" → advances past "DO" to reach "FLUIDO"
        "do", "da", "de", "del",
    }
    PUNCT_ONLY = re.compile(r"^[:\-_/|.,;]+$")

    label_y: float | None = None
    label_x_max: float    = 0.0
    label_x_start: float  = 9999.0
    for w in words:
        poly = w.polygon
        if not poly:
            continue
        wn = _norm(w.content)
        if any(wn == t or wn.startswith(t) for t in label_triggers):
            ys  = [p.y for p in poly]
            xs  = [p.x for p in poly]
            cy  = sum(ys) / len(ys)
            xr  = max(xs)
            xl  = min(xs)
            if label_y is None or cy < label_y:
                label_y       = cy
                label_x_max   = xr
                label_x_start = xl

    if label_y is None:
        return ""

    # Pass 2: extend label_x_max to cover ALL trigger words on the same line.
    # Example: "Frota/TAG" may be split by OCR into "Frota" + "TAG".
    # "TAG" is in FORM_KEYWORDS and would prematurely stop the value scan
    # unless we advance label_x_max past it first.
    for w in words:
        poly = w.polygon
        if not poly:
            continue
        wn = _norm(w.content)
        if any(wn == t or wn.startswith(t) for t in label_triggers):
            ys = [p.y for p in poly]
            xs = [p.x for p in poly]
            cy = sum(ys) / len(ys)
            xr = max(xs)
            if abs(cy - label_y) <= 0.10:   # same line tolerance
                label_x_max = max(label_x_max, xr)

    # Pass 3: crawl rightward through consecutive label-compound FORM_KEYWORDS.
    # Example: "Vol." → "fluido" → "adic." are all FORM_KEYWORDS that form the
    # multi-word label "Vol. fluido adic." — we must advance label_x_max past
    # all of them so the value scan starts only after the last label word.
    # We sort by left-edge and keep advancing as long as the next word:
    #   (a) is on the same line,
    #   (b) starts within X_GAP of the current label_x_max, AND
    #   (c) is itself a FORM_KEYWORD (i.e. part of the label, not the value).
    _sorted_words = sorted(
        [w for w in words if w.polygon],
        key=lambda w2: min(p.x for p in w2.polygon),
    )
    Y_TOL = 0.08
    X_GAP = 0.12
    changed = True
    while changed:
        changed = False
        for w in _sorted_words:
            poly = w.polygon
            xs   = [p.x for p in poly]
            ys   = [p.y for p in poly]
            cy   = sum(ys) / len(ys)
            xl   = min(xs)
            xr   = max(xs)
            wn   = _norm(w.content)
            if abs(cy - label_y) > Y_TOL:
                continue
            if xl <= label_x_max:          # already covered
                continue
            if xl > label_x_max + X_GAP:  # too far — value territory
                continue
            is_label_kw = (
                wn in FORM_KEYWORDS
                or wn.rstrip(":") in FORM_KEYWORDS
                or any(wn.startswith(k) for k in FORM_KEYWORDS if len(k) >= 4)
                or PUNCT_ONLY.match(w.content.strip())
            )
            if is_label_kw:
                label_x_max = max(label_x_max, xr)
                changed = True

    candidates: list[tuple[float, str]] = []
    for w in words:
        poly = w.polygon
        if not poly:
            continue
        xs = [p.x for p in poly]
        ys = [p.y for p in poly]
        cx = sum(xs) / len(xs)
        cy = sum(ys) / len(ys)
        if abs(cy - label_y) > Y_TOL:
            continue
        if cx <= label_x_max + X_GAP:
            continue
        candidates.append((cx, w.content))

    def _below_search() -> str:
        """
        Look for the first non-label word directly below the trigger label.
        Uses X-range filtering so we only pick up values within the label's
        horizontal span (avoids grabbing values from adjacent columns).
        Runs when horizontal scan found nothing — either because candidates was
        empty OR because all candidates were FORM_KEYWORDS that stopped the scan.
        """
        bw: list[tuple[float, float, str]] = []
        for w2 in words:
            p2 = w2.polygon
            if not p2:
                continue
            xs2 = [p.x for p in p2]
            ys2 = [p.y for p in p2]
            cx2 = sum(xs2) / len(xs2)
            cy2 = sum(ys2) / len(ys2)
            if cy2 <= label_y + 0.03 or cy2 > label_y + 0.80:
                continue
            # X filter: stay within the label's horizontal span (+ 0.5" right margin).
            # label_x_start = left edge of the first trigger word.
            # label_x_max   = right edge of the label after Pass 3.
            # This prevents grabbing values from adjacent columns on the same form row.
            if cx2 < label_x_start or cx2 > label_x_max + 0.50:
                continue
            wn2 = _norm(w2.content)
            if not wn2 or PUNCT_ONLY.match(w2.content.strip()):
                continue
            # Skip label words — they belong to adjacent field labels, not values
            if (wn2 in FORM_KEYWORDS
                    or wn2.rstrip(":") in FORM_KEYWORDS
                    or any(wn2.startswith(k) for k in FORM_KEYWORDS if len(k) >= 4)):
                continue
            bw.append((cy2, cx2, w2.content))
        if not bw:
            return ""
        bw.sort()
        first_cy2 = bw[0][0]
        line2 = [(cx2, c) for cy2, cx2, c in bw if abs(cy2 - first_cy2) <= Y_TOL]
        line2.sort()
        return " ".join(c for _, c in line2[:4]).strip()

    if not candidates:
        if search_below:
            return _below_search()
        return ""

    candidates.sort(key=lambda x: x[0])

    value_parts: list[str] = []
    for _, content in candidates:
        wn = _norm(content)
        if not wn or len(wn) < 2:
            # Melhoria 1: tokens curtos como "h" e "km" são unidades válidas —
            # anexa ao token anterior em vez de descartar
            if value_parts and re.match(r"^(h|km|kms|hrs?)$", content.strip().lower()):
                value_parts[-1] = value_parts[-1] + content.strip()
                break
            continue
        if PUNCT_ONLY.match(content.strip()):
            continue
        if re.match(r"^\d{7,}$", content.strip()):
            continue
        if wn in FORM_KEYWORDS:
            break
        if wn.rstrip(":") in FORM_KEYWORDS:
            break
        if any(wn.startswith(k) for k in FORM_KEYWORDS if len(k) >= 4):
            break
        value_parts.append(content)
        if len(value_parts) >= 4:
            break

    result = " ".join(value_parts).strip()
    if result:
        return result

    # Horizontal candidates existed but all were FORM_KEYWORDS (e.g. "Fluido"
    # from "FLUIDO TROCADO?" adjacent label) — the scan stopped immediately.
    # Fall through to below-label search before giving up.
    if search_below:
        return _below_search()
    return ""


def _read_field_easyocr(page_bytes: bytes, page, label_triggers: list[str],
                         extend_below: bool = False) -> str:
    """
    Stage 2: EasyOCR fallback — render page to image, crop the region
    to the right of the label, and run EasyOCR on it.
    When extend_below=True, also includes 0.5" below the label in the crop
    (for fields where the value is written on the line under the label).
    """
    reader = _get_easyocr_reader()
    if reader is None:
        return ""

    try:
        import fitz
        import numpy as np
        from PIL import Image as PILImage
    except ImportError:
        return ""

    words = getattr(page, "words", None) or []
    pw    = getattr(page, "width",  8.5)  or 8.5
    ph    = getattr(page, "height", 11.0) or 11.0

    label_y: float | None = None
    label_x_max: float    = 0.0
    label_x_start: float  = 9999.0
    for w in words:
        poly = w.polygon
        if not poly:
            continue
        wn = _norm(w.content)
        if any(t in wn for t in label_triggers):
            ys = [p.y for p in poly]
            xs = [p.x for p in poly]
            cy = sum(ys) / len(ys)
            xr = max(xs)
            xl = min(xs)
            if label_y is None or cy < label_y:
                label_y       = cy
                label_x_max   = xr
                label_x_start = xl

    if label_y is None:
        return ""

    # Extend label_x_max to cover ALL trigger words on the same line
    # (same fix as Stage 1: avoids cropping from before "TAG" in "Frota/TAG")
    for w in words:
        poly = w.polygon
        if not poly:
            continue
        wn = _norm(w.content)
        if any(t in wn for t in label_triggers):
            ys = [p.y for p in poly]
            xs = [p.x for p in poly]
            cy = sum(ys) / len(ys)
            xr = max(xs)
            if abs(cy - label_y) <= 0.10:
                label_x_max = max(label_x_max, xr)

    try:
        doc       = fitz.open(stream=page_bytes, filetype="pdf")
        fitz_page = doc[0]
        dpi       = 300
        mat       = fitz.Matrix(dpi / 72, dpi / 72)
        pix       = fitz_page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
        img       = PILImage.frombytes("L", [pix.width, pix.height], pix.samples)
        doc.close()
    except Exception:
        return ""

    scale_x = pix.width  / pw
    scale_y = pix.height / ph

    MARGIN_Y       = 0.14
    CROP_W         = 2.5
    BELOW_EXTEND   = 0.55  # extra inches below the label when extend_below=True
    if extend_below:
        # Crop from the label's left edge (value box is below the label)
        x1 = int(max(0, (label_x_start - 0.1) * scale_x))
        x2 = int(min(pix.width, (label_x_start + CROP_W) * scale_x))
        y1 = int(max(0, (label_y - MARGIN_Y) * scale_y))
        y2 = int(min(pix.height, (label_y + BELOW_EXTEND) * scale_y))
    else:
        x1 = int((label_x_max + 0.10) * scale_x)
        y1 = int(max(0, (label_y - MARGIN_Y) * scale_y))
        x2 = int(min(pix.width,  (label_x_max + 0.10 + CROP_W) * scale_x))
        y2 = int(min(pix.height, (label_y + MARGIN_Y) * scale_y))

    if x2 <= x1 or y2 <= y1 or (x2 - x1) < 20:
        return ""

    crop = img.crop((x1, y1, x2, y2))
    crop = crop.resize((crop.width * 2, crop.height * 2), PILImage.LANCZOS)
    arr  = np.array(crop)

    try:
        results  = reader.readtext(arr, detail=1, paragraph=False)
        MIN_CONF = 0.45
        parts    = [text for _, text, conf in results
                    if conf >= MIN_CONF and len(text.strip()) >= 2]
        text = " ".join(parts).strip()
        if re.match(r"^[\-_\s]+$", text):
            return ""
        return text
    except Exception:
        return ""


def _fallback_read_field(page_bytes: bytes | None, page,
                          label_triggers: list[str],
                          search_below: bool = False) -> str:
    result = _read_field_from_words(page, label_triggers, search_below=search_below)
    if result:
        return result
    if page_bytes is not None:
        result = _read_field_easyocr(page_bytes, page, label_triggers,
                                      extend_below=search_below)
    return result


# ── OIL DESCRIPTION NORMALIZER ────────────────────────────────────────────────
KNOWN_OIL_NAMES: list[str] = [
    "Texaco", "Petrobras", "Ipiranga", "Volvo", "John Deere",
    "Lubrax", "Lubrax API CI-4/SL", "Ipiranga ODM", "Cool Gard",
    "Shell", "Shell Spirax", "Castrol", "Mobil", "Total", "Motul",
    "15W40", "10W40", "5W30", "5W40", "20W50",
    "85W140", "75W80", "75W90", "80W90",
    "SAE 30", "SAE 40", "SAE 50",
    "1000 Marins 40",
]
_OIL_KNOWN_LOWER = {n.lower(): n for n in KNOWN_OIL_NAMES}

# OCR char substitution for viscosity prefix (before the W)
_VIS_OCR_PREFIX = {"I": "1", "i": "1", "l": "1", "L": "1", "O": "0", "o": "0", "S": "5", "s": "5"}
# OCR char substitution for viscosity suffix (after the W — digits only)
# Extended to cover common misreads seen in handwritten Brazilian forms:
#   g/G → 4  (handwritten 4 looks like g)
#   n/N → 4  (handwritten 4 open-top looks like n)
#   a/A → 0  (poorly-formed 0 looks like a)
#   H/h → 4  (already existed)
_VIS_OCR_SUFFIX = {
    "O": "0", "o": "0",           # 0 variants
    "I": "1", "i": "1", "l": "1", # 1 variants
    "H": "4", "h": "4",           # 4 variants
    "g": "4", "G": "4",           # 4 variants (g looks like 4)
    "n": "4", "N": "4",           # 4 variants (n open-top looks like 4)
    "a": "0", "A": "0",           # 0 variants (a can resemble 0)
}

# Words/tokens that are never part of the oil name
_OIL_NOISE = {
    "viscosidade", "viscosity", "de", "do", "da", "e", "para", "com",
    "un", "tudo", "que", "comeca", "pe", "provavelmente",
}

# Brands detected by prefix rule
# NOTE: more specific prefixes (e.g. "SPIRAX") must come BEFORE their parent
# ("SHELL") so the longer match wins first.
_BRAND_PREFIXES: list[tuple[str, str]] = [
    ("PETRONAS", "Petronas"),   # must be before "PET" so it wins
    ("PET",     "Petrobras"),
    ("LUBR",    "Lubrax"),
    ("TEXACO",  "Texaco"),
    ("VOLVO",   "Volvo"),
    ("IPIR",    "Ipiranga"),
    ("JOHN",    "John Deere"),
    ("COOL",    "Cool Gard"),
    ("SPIRA",   "Shell Spirax"),   # Shell Spirax (transmission/gear oil) — before SHELL
    ("SHELL",   "Shell"),
    ("CASTROL", "Castrol"),
    ("MOBIL",   "Mobil"),
    ("TOTAL",   "Total"),
    ("MOTUL",   "Motul"),
]


def _parse_viscosity(token: str) -> str | None:
    t = token.strip()

    # Extra leading digit: 115W40 → 15W40
    m = re.match(r"^1(\d{2}[Ww]\d{2,3})$", t)
    if m:
        result = _parse_viscosity(m.group(1))
        if result:
            return result

    # Already clean: 15W40
    m = re.match(r"^(\d{1,3})[Ww](\d{2,3})$", t)
    if m:
        return f"{m.group(1)}W{m.group(2)}"

    # OCR letter substitutions: ISW40, IOW40, 15l40, 15u90, 15v40, 15mno, 15uga …
    # W-separator misreads: l/L/u/U/v/V/m/M — handwritten W looks like these
    # Suffix misreads handled by _VIS_OCR_SUFFIX (g→4, n→4, a→0, etc.)
    m = re.match(r"^([A-Za-z0-9]{1,3})[WwlLuUvVmM]{1,2}([A-Za-z0-9]{2,3})$", t)
    if m:
        prefix = "".join(_VIS_OCR_PREFIX.get(c, c) for c in m.group(1))
        suffix = "".join(_VIS_OCR_SUFFIX.get(c, c) for c in m.group(2))
        if re.match(r"^\d+$", prefix) and re.match(r"^\d+$", suffix):
            grade = _fix_nonstandard_viscosity(f"{prefix}W{suffix}")
            # Extra OCR letter inflated prefix: "L15W40" → prefix="115" → try "15W40"
            if len(prefix) == 3 and prefix[0] == "1":
                stripped = _fix_nonstandard_viscosity(f"{prefix[1:]}W{suffix}")
                if stripped in _SAE_STANDARD:
                    return stripped
            return grade

    # Missing W: 15040 → 15W40
    m = re.match(r"^(\d{1,3})0(\d{2})$", t)
    if m and int(m.group(2)) in (30, 40, 50, 80, 90, 140):
        return f"{m.group(1)}W{m.group(2)}"

    # SAE mono-grade gear/engine oils (no W): 30, 40, 50, 80, 90, 140
    # Handles "90", "SAE90", "SAE 90" (after space-collapse in caller)
    m = re.match(r"^(?:SAE\s*)?(\d{2,3})$", t, re.I)
    if m:
        n = int(m.group(1))
        if n in (30, 40, 50, 80, 90, 140):
            return f"SAE {n}"

    return None


# SAE standard multigrade grades
_SAE_STANDARD = {
    "0W20", "5W20", "5W30", "5W40", "10W30", "10W40", "15W40", "20W50",
    "75W80", "75W90", "80W90", "85W90", "85W140",
    "SAE 30", "SAE 40", "SAE 50",
}
_SAE_CORRECT_PREFIX = {
    "0":  "0W20",  "5":  "5W40",  "10": "10W40", "15": "15W40",
    "19": "15W40", "20": "20W50",
    "75": "75W90", "80": "80W90", "85": "85W140",
}


def _fix_nonstandard_viscosity(grade: str) -> str:
    """Map non-standard SAE grades (OCR errors) to the nearest real grade."""
    if grade in _SAE_STANDARD:
        return grade
    m = re.match(r"^(\d+)W(\d+)$", grade)
    if not m:
        return grade
    corrected = _SAE_CORRECT_PREFIX.get(m.group(1))
    return corrected if corrected else grade


# ============================================================
# OIL REFERENCE CATALOG (fabricante/modelo/viscosidade) — used to fuzzy-correct
# OCR errors in those 3 fields and flag values that aren't in the catalog.
# ============================================================
_OIL_REF_PATH = os.path.join(os.path.dirname(__file__), "oil_reference.csv")
_OIL_REF_NULLS = {"", "-", "null", "none", "nan"}


@st.cache_data()
def _load_oil_reference():
    """Returns 3 dicts: normalized_value -> canonical_value, one per field."""
    import csv as _csv

    fab_map: dict[str, str] = {}
    mod_map: dict[str, str] = {}
    visc_map: dict[str, str] = {}

    def _visc_key_and_canon(raw: str) -> tuple[str, str]:
        s = raw.strip()
        s_wo_sae = re.sub(r"(?i)^sae\s+", "", s).strip()
        if re.match(r"^\d{1,3}[Ww]\d{0,3}$", s_wo_sae):
            canon = s_wo_sae.upper()
            return _norm(canon), canon
        return _norm(s), s

    try:
        with open(_OIL_REF_PATH, encoding="utf-8-sig") as f:
            for row in _csv.DictReader(f):
                fab  = (row.get("fabricante_oleo") or "").strip()
                mod  = (row.get("modelo_oleo") or "").strip()
                visc = (row.get("viscosidade") or "").strip()

                if fab and _norm(fab) not in _OIL_REF_NULLS:
                    k = _norm(fab)
                    fab_map.setdefault(k, fab)
                if mod and _norm(mod) not in _OIL_REF_NULLS:
                    k = _norm(mod)
                    mod_map.setdefault(k, mod)
                if visc and _norm(visc) not in _OIL_REF_NULLS:
                    k, canon = _visc_key_and_canon(visc)
                    if k:
                        visc_map.setdefault(k, canon)
    except FileNotFoundError:
        pass

    return fab_map, mod_map, visc_map


def _fuzzy_match_reference(value: str, ref_map: dict, cutoff: float = 0.78,
                            strip_sae: bool = False,
                            strip_trailing_visc: bool = False) -> tuple[str, bool]:
    """Returns (possibly-corrected value, matched). matched=False means the
    value wasn't found in the catalog (exact or close) — flag as suspicious.

    strip_trailing_visc: for model names, when the first pass fails, retry after
    removing a viscosity number glued onto the end of the name (e.g. the OCR read
    "Rimula Rf4C 154140" — the "154140" is a mis-read 15W40 stuck to the model).
    Catalog model names never carry the viscosity, so "Rimula Rf4C" then matches
    "RIMULA R4" cleanly."""
    if not value or not ref_map:
        return value, True
    if strip_sae:
        key = _norm(re.sub(r"(?i)^sae\s+", "", value).strip())
    else:
        key = _norm(value)
    if not key:
        return value, True
    if key in ref_map:
        return ref_map[key], True
    keys = list(ref_map.keys())
    best = get_close_matches(key, keys, n=1, cutoff=cutoff)
    if best:
        return ref_map[best[0]], True

    # Second pass: drop a trailing all-numeric / viscosity-shaped token and retry.
    if strip_trailing_visc:
        stripped = re.sub(r"\s+\d[\dWw]*$", "", value).strip()
        if stripped and stripped != value:
            key2 = _norm(stripped)
            if key2 and key2 != key:
                if key2 in ref_map:
                    return ref_map[key2], True
                best2 = get_close_matches(key2, keys, n=1, cutoff=cutoff)
                if best2:
                    return ref_map[best2[0]], True

    return value, False


def _normalize_oil_description(text: str) -> str:
    from difflib import get_close_matches

    if not text or not text.strip():
        return text

    # ── Pre-processing ────────────────────────────────────────────────────
    # Strip degree/ordinal/euro symbols — OCR artifacts in viscosity fields ("40°")
    # Azure sometimes returns € (U+20AC) instead of ° (U+00B0) for degree signs.
    text = re.sub(r"[°º€]", "", text)

    # Fix OCR-split / OCR-mangled brand names BEFORE token analysis
    # "SHoll" / "Sholl" / "SHell" variants → "Shell"
    text = re.sub(r"\bSH[oOeE][lL]{1,2}\b", "Shell", text)
    # "S Hell" / "S Holl" (brand split into two tokens) → "Shell"
    text = re.sub(r"\bS\s+[Hh][eEoO][lL]{1,2}\b", "Shell", text)
    # "I piranga" → "Ipiranga"
    text = re.sub(r"\bI\s+piranga\b", "Ipiranga", text, flags=re.I)

    # "15 W 40" / "15W 40" → "15W40"
    text = re.sub(r"(?<![A-Za-z])(\d{1,3})\s*[Ww]\s*(\d{2,3})(?!\d)", r"\1W\2", text)
    # "IS W40" (OCR split)
    text = re.sub(
        r"(?<!\w)([A-Za-z]{1,3})\s+[Ww]([A-Za-z0-9]{2,3})(?!\w)",
        lambda m: _parse_viscosity(m.group(1) + "W" + m.group(2)) or m.group(0),
        text,
    )
    # "15 UN 40"
    text = re.sub(r"(?<!\d)(\d{1,3})\s+[Uu][Nn]\s+(\d{2,3})(?!\d)", r"\1W\2", text)
    # "15 00 40"
    text = re.sub(r"(?<!\d)(\d{1,2})\s+00\s+(\d{2})(?!\d)", r"\1W\2", text)
    # "SW3"/"SW 3"/"SW30" → "5W30"; "SW4"/"SW40" → "5W40"
    text = re.sub(r"(?<![A-Za-z0-9])SW\s*3(?:0)?(?![A-Za-z0-9])", "5W30", text, flags=re.I)
    text = re.sub(r"(?<![A-Za-z0-9])SW\s*4(?:0)?(?![A-Za-z0-9])", "5W40", text, flags=re.I)

    # ── Whole-string synonyms ─────────────────────────────────────────────
    key = re.sub(r"\s+", " ", text.strip().lower())
    _SYNONYMS_INLINE = {
        "sw 3": "5W30",  "sw3": "5W30",
        "iow40": "10W40", "i0w40": "10W40", "l0w40": "10W40",
        "low40": "10W40", "1ow40": "10W40",
        # "lowago" = "10W40" with OCR garble (l→1, o→0, w→W, a→4 or noise, go→0)
        "lowago": "10W40", "lowago": "10W40",
        "sw30": "5W30",  "sw40": "5W40",
        "15w-40": "15W40", "85w-140": "85W140", "75w-80": "75W80",
        "lubrax apici-4/sl":  "Lubrax API CI-4/SL",
        "lubrax api ci4/sl":  "Lubrax API CI-4/SL",
        "lubrax apici4sl":    "Lubrax API CI-4/SL",
        "lubrac api ci-4/sl": "Lubrax API CI-4/SL",
        "lubrac api ci -4/sl":"Lubrax API CI-4/SL",
        "lubrax api ci-4 sl": "Lubrax API CI-4/SL",
        "coolgard": "Cool Gard", "cool guard": "Cool Gard",
        "is who": "15W40", "iswho": "15W40",
        # Non-standard OCR grades
        "19w40": "15W40", "19W40": "15W40",
        "l5w40": "15W40", "l5W40": "15W40",
        "i5w40": "15W40", "i5W40": "15W40",
        # Combined OCR-corrupted brand+grade strings
        "shell 19w40": "Shell 15W40",
        # "1000 Marins 40" product name — "MARING" = OCR misread of "Marins"
        "1000 maring 40": "1000 Marins 40",
        "1000 marins 40": "1000 Marins 40",
        "texaco 1000 maring 40": "Texaco 1000 Marins 40",
        "texaco 1000 marins 40": "Texaco 1000 Marins 40",
        "shell 19ugo": "Shell 15W40",
        "sholl 19 ugo": "Shell 15W40",
        "s hell 15 mno": "Shell 15W40",
        "s hell 15w40": "Shell 15W40",
    }
    if key in _SYNONYMS_INLINE:
        return _SYNONYMS_INLINE[key]

    # ── Token-by-token analysis ───────────────────────────────────────────
    brand: str | None     = None
    viscosity: str | None = None
    lubrax_api            = False

    tokens = text.split()
    skip_next = False
    for i, tok in enumerate(tokens):
        if skip_next:
            skip_next = False
            continue

        tok_alpha = re.sub(r"[^A-Za-z0-9/\-]", "", tok)
        if not tok_alpha:
            continue
        tok_up = tok_alpha.upper()
        tok_lo = tok_alpha.lower()

        if tok_lo in _OIL_NOISE:
            continue

        # "SAE N" two-token mono-grade (e.g. "SAE 90"): look one token ahead
        if tok_up == "SAE" and i + 1 < len(tokens):
            next_alpha = re.sub(r"[^A-Za-z0-9]", "", tokens[i + 1])
            combined   = f"SAE{next_alpha}"   # e.g. "SAE90"
            vis = _parse_viscosity(combined)
            if vis:
                viscosity  = vis
                skip_next  = True
                continue

        # Try viscosity BEFORE long-digit filter so "15040"→"15W40" works
        vis = _parse_viscosity(tok_alpha)
        if vis:
            viscosity = _fix_nonstandard_viscosity(vis)
            continue

        # Two-token viscosity lookahead: "15 mno" → "15mno" → "15W40"
        # Handles OCR where W and suffix digits are garbled into adjacent tokens
        # Guard: only accept results that are valid/correctable SAE grades
        if i + 1 < len(tokens):
            next_alpha2 = re.sub(r"[^A-Za-z0-9]", "", tokens[i + 1])
            combined2   = tok_alpha + next_alpha2
            vis2 = _parse_viscosity(combined2)
            if vis2:
                vis2_fixed = _fix_nonstandard_viscosity(vis2)
                # Only accept if the result is a known SAE standard grade
                if vis2_fixed in _SAE_STANDARD or re.match(r"^SAE \d+$", vis2_fixed):
                    viscosity  = vis2_fixed
                    skip_next  = True
                    continue

        # Skip barcodes / horimeter numbers
        if re.match(r"^\d{5,}$", tok_alpha):
            continue

        # Brand prefix rules
        matched_brand = None
        for prefix, canonical in _BRAND_PREFIXES:
            if tok_up.startswith(prefix):
                matched_brand = canonical
                break

        if matched_brand:
            rest = " ".join(tokens[i + 1:]).upper()
            if matched_brand == "Lubrax":
                if re.search(r"API|CI.?4", rest):
                    matched_brand = "Lubrax API CI-4/SL"
                    lubrax_api    = True
            elif matched_brand == "Ipiranga":
                if "ODM" in rest:
                    matched_brand = "Ipiranga ODM"
            brand = matched_brand
            continue

        if lubrax_api and tok_up in ("API", "CI", "SL"):
            continue
        if lubrax_api and re.match(r"^CI.?4", tok_up):
            continue

    # ── Build result ──────────────────────────────────────────────────────
    parts = []
    if brand:
        parts.append(brand)
    if viscosity:
        parts.append(viscosity)

    if parts:
        return " ".join(parts)

    # Fuzzy whole-string match
    match = get_close_matches(text.lower(), list(_OIL_KNOWN_LOWER.keys()), n=1, cutoff=0.78)
    if match:
        return _OIL_KNOWN_LOWER[match[0]]

    # Nothing recognized in our known list → keep original value as-is.
    # The field mapping now guarantees we're only receiving FLUIDO-section values,
    # so unknown values are still valid oil descriptions.
    return text


def _join_compartments(labels: list) -> str:
    """Deduplicate, sort by form order (COMPARTMENT_LABELS), join with '/'.
    'Ponto de Coleta' is the section header, never a valid compartment value —
    it is always filtered out so the read-below fallback can resolve the real name."""
    order  = {lbl: i for i, lbl in enumerate(COMPARTMENT_LABELS)}
    unique = [l for l in dict.fromkeys(labels) if l != "Ponto de Coleta"]
    unique.sort(key=lambda l: order.get(l, 999))
    return "/".join(unique)


def _find_compartment_via_page_marks(page, page_bytes: bytes | None = None) -> str:
    """Match each Azure 'selected' mark to its compartment label using Euclidean
    distance to the label centre.  Optionally verifies the mark with pixel
    darkness to eliminate false positives where Azure marks an empty circle."""

    ph = getattr(page, "height", 11.0) or 11.0
    pw = getattr(page, "width",   8.5) or 8.5
    COMP_X_MIN = pw * 0.55
    COMP_Y_MIN = ph * 0.06
    COMP_Y_MAX = ph * 0.94
    MIN_CONF   = 0.70
    MAX_DIST   = 0.40   # max inches between mark centre and label x_min
    PIXEL_MIN  = 40     # minimum inner-ring darkness score to accept a mark

    # ── Build pixel image for verification (if page_bytes provided) ───────────
    img_arr   = None
    scale_x   = scale_y = 1.0
    w_px = h_px = 0
    if page_bytes:
        try:
            import fitz, numpy as np
            from PIL import Image
            import io as _io
            _doc = fitz.open(stream=page_bytes, filetype="pdf")
            _pix = _doc[0].get_pixmap(dpi=200, alpha=False)
            _doc.close()
            _img     = Image.open(_io.BytesIO(_pix.tobytes("png"))).convert("L")
            img_arr  = np.array(_img, dtype=np.float32)
            h_px, w_px = img_arr.shape
            scale_x  = w_px / pw
            scale_y  = h_px / ph
        except Exception:
            img_arr = None

    # Build word_list for _find_label_positions
    word_list = []
    for w in (getattr(page, "words", None) or []):
        poly = w.polygon
        if not poly:
            continue
        xs = [p.x for p in poly]
        ys = [p.y for p in poly]
        cx = sum(xs) / len(xs)
        cy = sum(ys) / len(ys)
        if cx < COMP_X_MIN or cy < COMP_Y_MIN or cy > COMP_Y_MAX:
            continue
        word_list.append({
            "norm":  _norm(w.content),
            "cx":    cx,
            "cy":    cy,
            "x_min": min(xs),
            "x_max": max(xs),
        })

    label_pos = _find_label_positions(word_list)
    if _PM_DEBUG:
        eixo_labels = {k: v for k, v in label_pos.items() if "ixo" in k.lower() or "raxa" in k.lower()}
        print(f"  [PM_DBG] label_pos eixo/graxa: {eixo_labels}")
    if not label_pos:
        return ""

    candidates: list[tuple[float, float, str]] = []   # (conf, dist, canon)
    for sm in (getattr(page, "selection_marks", None) or []):
        if sm.state != "selected" or sm.confidence < MIN_CONF:
            continue
        poly = sm.polygon
        if not poly:
            continue
        scx = sum(p.x for p in poly) / len(poly)
        scy = sum(p.y for p in poly) / len(poly)
        if scx < COMP_X_MIN or scy < COMP_Y_MIN or scy > COMP_Y_MAX:
            continue

        # Find closest label (mark must be left of label x_min).
        # Initial match uses the raw mark position.  For filled circles the
        # position is later corrected via upward scan (see pixel verification
        # block below) and the label re-matched with the corrected Y.
        best_label, best_dist = None, float("inf")
        second_dist_raw = float("inf")
        _all_dists: list[tuple[float, str]] = []
        for label, (lx, ly) in label_pos.items():
            if scx >= lx:
                continue
            dist = ((scx - lx) ** 2 + (scy - ly) ** 2) ** 0.5
            _all_dists.append((dist, label))
            if dist <= MAX_DIST:
                if dist < best_dist:
                    second_dist_raw = best_dist
                    best_dist       = dist
                    best_label      = label
                elif dist < second_dist_raw:
                    second_dist_raw = dist
        if _PM_DEBUG:
            _top3 = sorted(_all_dists)[:4]
            print(f"    [PM_DBG] mark scx={scx:.3f} scy={scy:.3f} -> raw_best={best_label!r}@{best_dist:.3f} 2nd@{second_dist_raw:.3f} | top4={_top3}")
        if best_label is None:
            continue
        # Require the winner to be clearly closer than the second label.
        # If two labels are virtually equidistant (ratio < 1.1), this mark is
        # at an ambiguous position and should not map to either label.
        if second_dist_raw / max(best_dist, 1e-6) < 1.10:
            if _PM_DEBUG:
                print(f"    [PM_DBG] RAW MATCH AMBIGUOUS (ratio {second_dist_raw/max(best_dist,1e-6):.2f} < 1.1), skipping")
            continue

        # Pixel verification with filled-circle centre correction.
        # Azure selection marks on FILLED circles are detected near the bottom of
        # the ink blob (up to ~0.20 in below actual centre).  To find the correct
        # label we: (1) reject marks with no ink (score < PIXEL_MIN), (2) for marks
        # with very high score (filled circle), scan upward to find the geometric
        # centre of the filled region and re-match using that corrected position.
        if img_arr is not None:
            data = _measure_circle_data([(scx, scy, "_")], img_arr, scale_x, scale_y, w_px, h_px)
            pix_score = data[0][0] if data else 0.0
            if _PM_DEBUG:
                print(f"    [PM_DBG] pix_score={pix_score:.1f} at ({scx:.3f},{scy:.3f})")
            if pix_score < PIXEL_MIN:
                if _PM_DEBUG:
                    print(f"    [PM_DBG] REJECTED (score {pix_score:.1f} < PIXEL_MIN {PIXEL_MIN})")
                continue

            FILLED_THRESH = 150   # score above this → filled circle, likely offset
            if pix_score >= FILLED_THRESH:
                # Scan upward from mark centre to find top of dark region
                r_scan  = 0.06      # half-width of measurement patch (inches)
                step    = 0.04      # step size when scanning (inches)
                y_probe = scy
                top_dark_y = scy    # last Y where patch was still dark
                # Limit scan to 0.20 in to avoid crossing into an adjacent circle's
                # printed outline and over-correcting the estimated centre.
                while y_probe > scy - 0.20:
                    y_probe -= step
                    d2 = _measure_circle_data([(scx, y_probe, "_")], img_arr, scale_x, scale_y, w_px, h_px)
                    s2 = d2[0][0] if d2 else 0.0
                    if s2 >= PIXEL_MIN:
                        top_dark_y = y_probe
                    else:
                        break
                # Estimated centre = midpoint of detected dark region
                scy_corrected = (top_dark_y + scy) / 2.0

                # Re-match label using corrected Y; also track second-best distance
                best_label_c, best_dist_c = None, float("inf")
                second_dist_c = float("inf")
                for label_c, (lx_c, ly_c) in label_pos.items():
                    if scx >= lx_c:
                        continue
                    dist_c = ((scx - lx_c) ** 2 + (scy_corrected - ly_c) ** 2) ** 0.5
                    if dist_c <= MAX_DIST:
                        if dist_c < best_dist_c:
                            second_dist_c = best_dist_c
                            best_dist_c   = dist_c
                            best_label_c  = label_c
                        elif dist_c < second_dist_c:
                            second_dist_c = dist_c
                # Only accept correction when the winner is clearly unambiguous:
                # second-best must be at least 1.5× farther than the winner.
                # If ambiguous, keep the original (uncorrected) label match.
                if best_label_c is not None:
                    ratio = second_dist_c / max(best_dist_c, 1e-6)
                    if _PM_DEBUG:
                        print(f"    [PM_DBG] corr scy={scy_corrected:.3f} top_dark={top_dark_y:.3f} "
                              f"corrected_best={best_label_c!r}@{best_dist_c:.3f} "
                              f"second_dist={second_dist_c:.3f} ratio={ratio:.2f}")
                    if ratio >= 1.5:
                        best_label = best_label_c
                        best_dist  = best_dist_c
                    elif _PM_DEBUG:
                        print(f"    [PM_DBG] correction REJECTED (ratio {ratio:.2f} < 1.5), keeping {best_label!r}")

        if _PM_DEBUG:
            print(f"    [PM_DBG] FINAL label={best_label!r}")
        canon = COMPARTMENT_OPTIONS.get(_norm(best_label), best_label)
        candidates.append((sm.confidence, best_dist, canon))

    if not candidates:
        return ""

    candidates.sort(key=lambda x: (-x[0], x[1]))

    # Móvel/Industrial filtering: if specific compartments are also detected,
    # drop the category label and return the specific one.
    all_canons  = [c for _, _, c in candidates]
    specific    = [c for c in all_canons if c not in CATEGORY_LABELS]
    if specific:
        for _conf, _dist, canon in candidates:
            if canon not in CATEGORY_LABELS:
                return canon

    top_conf, top_dist, top_canon = candidates[0]
    if len(candidates) >= 2:
        sec_conf, sec_dist, _ = candidates[1]
        if top_conf == sec_conf and abs(top_dist - sec_dist) < 0.03:
            return ""
    return top_canon


def _find_oleo_trocado_via_page_marks(page) -> str:
    ph = getattr(page, "height", 8.0)  or 8.0
    pw = getattr(page, "width",  12.6) or 12.6

    SIM_X_MIN = pw * 0.38
    SIM_X_MAX = pw * 0.57
    SIM_Y_MIN = ph * 0.40
    SIM_Y_MAX = ph * 0.80
    MIN_CONF  = 0.70

    sim_nao = {"sim": "Sim", "nao": "Não", "si": "Sim", "no": "Não"}

    word_positions: list[tuple[float, float, str]] = []
    for w in (getattr(page, "words", None) or []):
        poly = w.polygon
        if poly:
            cx = sum(p.x for p in poly) / len(poly)
            cy = sum(p.y for p in poly) / len(poly)
            word_positions.append((cx, cy, w.content))

    candidates: list[tuple[float, str]] = []
    for sm in (getattr(page, "selection_marks", None) or []):
        if sm.state != "selected" or sm.confidence < MIN_CONF:
            continue
        poly = sm.polygon
        if not poly:
            continue
        scx = sum(p.x for p in poly) / len(poly)
        scy = sum(p.y for p in poly) / len(poly)

        if not (SIM_X_MIN <= scx <= SIM_X_MAX and SIM_Y_MIN <= scy <= SIM_Y_MAX):
            continue

        best_label, best_dist = None, float("inf")
        for wx, wy, wtext in word_positions:
            wn = _norm(wtext)
            if wn in sim_nao:
                d = ((wx - scx) ** 2 + (wy - scy) ** 2) ** 0.5
                if d < best_dist:
                    best_dist  = d
                    best_label = sim_nao[wn]

        if best_label:
            score = sm.confidence * 10 - best_dist
            candidates.append((score, best_label))

    if not candidates:
        return ""
    candidates.sort(reverse=True)
    return candidates[0][1]


def _find_oleo_trocado_via_pixel_detection(page_pdf_bytes: bytes, page) -> str:
    """
    Detect 'Óleo trocado' (Sim/Não) using label-first + inner/ring ratio pixel score,
    the same technique used for compartment detection.

    Flow:
      Phase 1 — Azure marks as circle positions:
        For each label (Sim / Não), find the Azure selection_mark with the smallest
        Y-distance that lies to the left of the label. Measure inner/ring darkness.
      Phase 2 — Fixed offset fallback:
        If no Azure marks were found, probe at a fixed offset to the left of each label.
    """
    import fitz
    import numpy as np
    from PIL import Image
    import io

    if not page_pdf_bytes or page is None:
        return ""

    try:
        doc     = fitz.open(stream=page_pdf_bytes, filetype="pdf")
        pix     = doc[0].get_pixmap(dpi=200, alpha=False)
        doc.close()
        img     = Image.open(io.BytesIO(pix.tobytes("png"))).convert("L")
        img_arr = np.array(img, dtype=np.float32)
        h_px, w_px = img_arr.shape

        # Use landscape defaults (same as _find_oleo_trocado_via_page_marks)
        ph = getattr(page, "height", 8.0)  or 8.0
        pw = getattr(page, "width",  12.6) or 12.6
        scale_x = w_px / pw
        scale_y = h_px / ph

        # Region where Sim/Não labels live (same bounds as the old function)
        SIM_X_MIN = pw * 0.30
        SIM_X_MAX = pw * 0.65
        SIM_Y_MIN = ph * 0.30
        SIM_Y_MAX = ph * 0.88

        _SIM_NAO = {"sim": "Sim", "si": "Sim", "nao": "Não", "no": "Não"}

        # ── Collect label positions ───────────────────────────────────────
        label_pos: dict[str, tuple[float, float]] = {}   # "Sim"/"Não" → (cx, cy)
        for w in (getattr(page, "words", None) or []):
            poly = w.polygon
            if not poly:
                continue
            xs  = [p.x for p in poly]
            ys  = [p.y for p in poly]
            cx  = sum(xs) / len(xs)
            cy  = sum(ys) / len(ys)
            wn  = _norm(w.content)
            if wn not in _SIM_NAO:
                continue
            if not (SIM_X_MIN <= cx <= SIM_X_MAX and SIM_Y_MIN <= cy <= SIM_Y_MAX):
                continue
            label = _SIM_NAO[wn]
            # Keep the first (topmost) occurrence of each label
            if label not in label_pos or cy < label_pos[label][1]:
                label_pos[label] = (cx, cy)

        if not label_pos:
            return ""

        # ── Ink-color scan (primary method) ─────────────────────────────
        # More robust than Azure's selection_marks: a handwritten X is sometimes
        # not registered as a "selected" mark at all by Azure's base OCR layout
        # (e.g. it only finds an unrelated low-confidence mark near the OTHER
        # label), which previously caused single-candidate false positives.
        # Scanning ink color directly at the circle position — independent of
        # whether Azure detected a mark glyph there — avoids that failure mode.
        img_rgb = np.array(Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB"), dtype=np.float32)
        INK_OFFSETS  = [0.10, 0.13, 0.16, 0.18, 0.20, 0.22, 0.25, 0.28]
        INK_THRESH   = 15.0
        ink_scores: dict[str, float] = {}
        for label, (lx, ly) in label_pos.items():
            best = 0.0
            for off in INK_OFFSETS:
                s = _ink_score_at(lx - off, ly, img_rgb, scale_x, scale_y, w_px, h_px)
                if s > best:
                    best = s
            ink_scores[label] = best
        if ink_scores:
            sorted_ink = sorted(ink_scores.items(), key=lambda x: -x[1])
            best_label, best_score = sorted_ink[0]
            second_score = sorted_ink[1][1] if len(sorted_ink) > 1 else 0.0
            if best_score >= INK_THRESH and (best_score - second_score) >= 10.0:
                return best_label

        # ── Collect Azure selection marks in the region ───────────────────
        marks: list[tuple[float, float]] = []
        for sm in (getattr(page, "selection_marks", None) or []):
            poly = sm.polygon
            if not poly:
                continue
            xs  = [p.x for p in poly]
            ys  = [p.y for p in poly]
            scx = sum(xs) / len(xs)
            scy = sum(ys) / len(ys)
            if SIM_X_MIN <= scx <= SIM_X_MAX and SIM_Y_MIN <= scy <= SIM_Y_MAX:
                marks.append((scx, scy))

        # ── Phase 1: exclusive mark assignment (Euclidean distance) ─────
        # Sim and Não are on the same row → Y-distance alone assigns both to
        # the same mark (identical scores, gap=0). We solve this with exclusive
        # assignment: each mark goes to its CLOSEST label; labels without a mark
        # fall back to a fixed offset so we still have a measurement for both.
        circle_positions: list[tuple[float, float, str]] = []
        Y_MARK = 0.22

        if marks:
            # Step 1: for each mark, find the closest label (must be to right of mark)
            mark_to_best: dict[tuple, tuple[str, float]] = {}
            for scx, scy in marks:
                for label, (lx, ly) in label_pos.items():
                    if scx >= lx - 0.02:
                        continue
                    dist = ((scx - lx) ** 2 + (scy - ly) ** 2) ** 0.5
                    if dist > Y_MARK * 2:
                        continue
                    prev = mark_to_best.get((scx, scy))
                    if prev is None or dist < prev[1]:
                        mark_to_best[(scx, scy)] = (label, dist)

            # Step 2: for each label, keep only the closest assigned mark
            label_to_mark: dict[str, tuple[float, float, float]] = {}
            for (mx, my), (label, dist) in mark_to_best.items():
                prev = label_to_mark.get(label)
                if prev is None or dist < prev[2]:
                    label_to_mark[label] = (mx, my, dist)

            # Step 3: only use labels that have an actual Azure mark assigned.
            # Do NOT fall back to estimated offsets here — with only 2 options
            # any dark area at an estimated position creates false positives.
            for label, (mx, my, _) in label_to_mark.items():
                circle_positions.append((mx, my, label))

        circle_data_p1 = _measure_circle_data(
            circle_positions, img_arr, scale_x, scale_y, w_px, h_px, r_in=0.08
        )

        if circle_data_p1:
            sorted_d  = sorted(circle_data_p1, key=lambda x: -x[0])
            best_s, best_lbl = sorted_d[0]
            second_s  = sorted_d[1][0] if len(sorted_d) > 1 else 0.0
            gap       = best_s - second_s
            if best_s >= 40.0 and gap >= 15.0:
                return best_lbl
            # gap too small → fall through to P2 (don't abort — with exclusive
            # assignment this only happens when both positions are truly ambiguous)

        # No Azure marks found in the region → pixel detection cannot help.
        # The old mark-state fallback (_find_oleo_trocado_via_page_marks) will
        # handle this case more conservatively (requires state=selected + conf>=0.70).
        return ""

    except Exception:
        return ""


def _read_ponto_de_coleta_description(page) -> str:
    ph = getattr(page, "height", 11.0) or 11.0
    pw = getattr(page, "width",  8.5)  or 8.5
    COMP_X_MIN = pw * 0.55

    words_in_zone = []
    for w in (getattr(page, "words", None) or []):
        poly = w.polygon
        if not poly:
            continue
        xs = [p.x for p in poly]
        ys = [p.y for p in poly]
        cx = sum(xs) / len(xs)
        cy = sum(ys) / len(ys)
        if cx < COMP_X_MIN:
            continue
        words_in_zone.append({"content": w.content, "norm": _norm(w.content), "cx": cx, "cy": cy})

    ponto_y = None
    for wi in words_in_zone:
        if "ponto" not in wi["norm"]:
            continue
        for wj in words_in_zone:
            if abs(wj["cy"] - wi["cy"]) < 0.12 and "coleta" in wj["norm"]:
                ponto_y = wi["cy"]
                break
        if ponto_y is not None:
            break

    if ponto_y is None:
        return ""

    BELOW_MIN = 0.08
    BELOW_MAX = 0.65

    FORM_LABELS = {
        "observacoes", "feedback", "observacoesfeedback",
        "observacoes/feedback", "observacoes/feedbac", "observacoesfeedbac",
        "ponto", "coleta",
        "data", "horimetro", "horometro", "km", "periodo",
        "comentario", "comentarios", "observacao", "observacoes:",
        "numero", "frasco", "chassi", "serie", "viscosidade",
        "descricao", "volume", "adicionado", "codigo", "externo",
    }

    # Norms of all compartment-list labels that are PRINTED on the form.
    # Words that match these norms within the checkbox-row spacing (dy < 0.35)
    # are printed form labels, not user-written descriptions — skip them.
    comp_norms = {_norm(lbl) for lbl in COMPARTMENT_LABELS}
    # Also individual tokens of multi-word labels (e.g. "final", "dianteiro")
    for lbl in COMPARTMENT_LABELS:
        for tok in _norm(lbl).split():
            if len(tok) >= 4:
                comp_norms.add(tok)

    below_words = []
    for w in words_in_zone:
        dy = w["cy"] - ponto_y
        if dy < BELOW_MIN or dy > BELOW_MAX:
            continue
        if w["norm"] in FORM_LABELS:
            continue
        if len(w["norm"]) < 2:
            continue
        # Skip printed compartment-list labels at any distance — they are always
        # part of the form's checkbox section, never user-written descriptions.
        if w["norm"] in comp_norms:
            continue
        below_words.append(w)

    if not below_words:
        return ""

    below_words.sort(key=lambda x: (round(x["cy"], 1), x["cx"]))

    lines: list[list[dict]] = []
    current_line = [below_words[0]]
    for w in below_words[1:]:
        if abs(w["cy"] - current_line[0]["cy"]) < 0.12:
            current_line.append(w)
        else:
            lines.append(current_line)
            current_line = [w]
    lines.append(current_line)

    first_line = sorted(lines[0], key=lambda x: x["cx"])
    text = " ".join(w["content"] for w in first_line).strip()
    return text if len(text) >= 2 else ""


def _pixel_score_at(x_in: float, y_in: float, page_bytes: bytes, pw: float, ph: float) -> float:
    """Return inner-ring darkness score at position (x_in, y_in) in page inches.
    Returns 0.0 on error or if image cannot be loaded."""
    try:
        import fitz, numpy as np
        from PIL import Image
        import io as _io
        _doc = fitz.open(stream=page_bytes, filetype="pdf")
        _pix = _doc[0].get_pixmap(dpi=200, alpha=False)
        _doc.close()
        _img    = Image.open(_io.BytesIO(_pix.tobytes("png"))).convert("L")
        arr     = np.array(_img, dtype=np.float32)
        h_px, w_px = arr.shape
        sx = w_px / pw
        sy = h_px / ph
        data = _measure_circle_data([(x_in, y_in, "_")], arr, sx, sy, w_px, h_px)
        return data[0][0] if data else 0.0
    except Exception:
        return 0.0


def _measure_circle_data(circle_positions, img_arr, scale_x, scale_y, w_px, h_px, r_in=0.09):
    import numpy as np
    results  = []
    r_inner  = r_in * 0.55
    r_outer  = r_in
    for cx, cy, label in circle_positions:
        xi1 = max(0, int((cx - r_inner) * scale_x))
        xi2 = min(w_px, int((cx + r_inner) * scale_x))
        yi1 = max(0, int((cy - r_inner) * scale_y))
        yi2 = min(h_px, int((cy + r_inner) * scale_y))
        xo1 = max(0, int((cx - r_outer) * scale_x))
        xo2 = min(w_px, int((cx + r_outer) * scale_x))
        yo1 = max(0, int((cy - r_outer) * scale_y))
        yo2 = min(h_px, int((cy + r_outer) * scale_y))
        if xi1 >= xi2 or yi1 >= yi2 or xo1 >= xo2 or yo1 >= yo2:
            continue
        inner      = img_arr[yi1:yi2, xi1:xi2]
        outer      = img_arr[yo1:yo2, xo1:xo2]
        inner_dark = float(255.0 - np.mean(inner))
        outer_dark = float(255.0 - np.mean(outer))
        ratio      = inner_dark / max(outer_dark, 5.0)
        score      = inner_dark * min(ratio, 2.0)
        results.append((score, label))
    return results


def _category_from_circle_data(circle_data, threshold: float = 70.0) -> str:
    """Return 'Móvel' or 'Industrial' if their pixel score >= threshold, else ''."""
    for score, label in sorted(circle_data, key=lambda x: -x[0]):
        canon = COMPARTMENT_OPTIONS.get(_norm(label), label)
        if canon in CATEGORY_LABELS and score >= threshold:
            return canon
    return ""


def _ink_score_at(cx: float, cy: float, img_rgb, scale_x: float, scale_y: float,
                   w_px: int, h_px: int, r_in: float = 0.09) -> float:
    """Score (0-100) of pen-ink density (blue or black) inside the circle at (cx, cy),
    robust to brownish/yellowish paper stains and dirt — which have low saturation
    in a different way than black ink, and lack the blue-channel boost of pen ink."""
    import numpy as np
    r_inner = r_in * 0.55
    xi1 = max(0, int((cx - r_inner) * scale_x))
    xi2 = min(w_px, int((cx + r_inner) * scale_x))
    yi1 = max(0, int((cy - r_inner) * scale_y))
    yi2 = min(h_px, int((cy + r_inner) * scale_y))
    if xi1 >= xi2 or yi1 >= yi2:
        return 0.0
    patch = img_rgb[yi1:yi2, xi1:xi2]
    r, g, b = patch[:, :, 0], patch[:, :, 1], patch[:, :, 2]

    # Blue pen ink: blue channel clearly dominant over red/green.
    blue_mask = (b - r > 12) & (b - g > 8)

    # Black pen ink: dark AND low saturation (paper stains tend to be brownish —
    # red channel notably higher than blue — so they don't pass this check).
    brightness = (r + g + b) / 3.0
    max_c = np.maximum(np.maximum(r, g), b)
    min_c = np.minimum(np.minimum(r, g), b)
    sat   = max_c - min_c
    black_mask = (brightness < 90) & (sat < 25)

    ink_mask = blue_mask | black_mask
    return float(np.mean(ink_mask)) * 100.0


def _category_from_ink(label_pos: dict, img_rgb, scale_x: float, scale_y: float,
                        w_px: int, h_px: int, comp_x_min: float, threshold: float = 8.0) -> str:
    """Scan the typical checkbox-circle offsets to the left of the Móvel/Industrial
    labels and score pen-ink color there. More robust than raw darkness scoring
    (_category_from_circle_data) because it ignores brownish paper stains/dirt that
    would otherwise look "dark" but aren't actually pen ink."""
    OFFSETS = [0.10, 0.13, 0.16, 0.18, 0.20, 0.22, 0.28, 0.34, 0.40]
    best_label, best_score = "", 0.0
    for label, (lx, ly) in label_pos.items():
        canon = COMPARTMENT_OPTIONS.get(_norm(label), label)
        if canon not in CATEGORY_LABELS:
            continue
        for off in OFFSETS:
            cx_try = lx - off
            if cx_try < comp_x_min:
                continue
            score = _ink_score_at(cx_try, ly, img_rgb, scale_x, scale_y, w_px, h_px)
            if score > best_score:
                best_score, best_label = score, canon
    return best_label if best_score >= threshold else ""


def _pick_winner(circle_data):
    if not circle_data:
        return ""
    sorted_data  = sorted(circle_data, key=lambda x: -x[0])
    scores       = [s for s, _ in circle_data]
    median_score = sorted(scores)[len(scores) // 2]
    best_score, best_raw_label = sorted_data[0]
    second_score = sorted_data[1][0] if len(sorted_data) > 1 else 0.0

    ABS_MIN    = 45.0
    REL_MARGIN = 20.0
    MIN_GAP    = 18.0
    if best_score < ABS_MIN:
        return ""
    if (best_score - median_score) < REL_MARGIN:
        return ""
    if (best_score - second_score) < MIN_GAP:
        # When two labels share the exact same mark (equidistant label positions on
        # the form, e.g. "Eixo Dianteiro / Eixo Traseiro" on one printed row), both
        # get assigned the same circle and measure identical scores.  If the score is
        # clearly above an empty printed circle (≥ 120), the circle IS filled —
        # return the first in sorted order (which follows COMPARTMENT_LABELS form order).
        if best_score >= 120.0 and abs(best_score - second_score) < 1.0:
            pass   # fall through to winner selection below
        else:
            return ""
    winner = COMPARTMENT_OPTIONS.get(_norm(best_raw_label), best_raw_label)
    # Móvel/Industrial filter: if a specific compartment also scores above ABS_MIN,
    # prefer it over the category label.
    if winner in CATEGORY_LABELS:
        for sc, lbl in sorted_data[1:]:
            if sc < ABS_MIN:
                break
            specific = COMPARTMENT_OPTIONS.get(_norm(lbl), lbl)
            if specific not in CATEGORY_LABELS:
                return specific
    return winner


def _pick_winners(circle_data):
    """Like _pick_winner but returns ALL labels that are clearly above background.
    Used when multiple compartments can be marked on the same form."""
    if not circle_data:
        return []
    sorted_data  = sorted(circle_data, key=lambda x: -x[0])
    scores       = [s for s, _ in circle_data]
    median_score = sorted(scores)[len(scores) // 2]

    ABS_MIN    = 45.0
    REL_MARGIN = 20.0

    winners = []
    for score, raw_label in sorted_data:
        if score < ABS_MIN:
            break
        if (score - median_score) < REL_MARGIN:
            continue
        canon = COMPARTMENT_OPTIONS.get(_norm(raw_label), raw_label)
        if canon not in winners:
            winners.append(canon)
    return winners


def _find_label_positions(word_list, Y_MATCH=0.13):
    positions = {}
    # Secondary words (e.g. "Dianteiro", "Traseiro") must be on the SAME text row
    # as their anchor. Using Y_MATCH here lets words from adjacent rows (spaced ~0.13)
    # falsely qualify, causing both "Eixo Dianteiro" and "Eixo Traseiro" to anchor on
    # the same "Eixo" word.  A tighter threshold keeps only true same-row words.
    SAME_ROW = Y_MATCH * 0.55  # ~0.07 in — strict same-row match

    for label in COMPARTMENT_LABELS:
        sig = [w for w in _norm(label).split() if len(w) >= 3]
        if not sig:
            continue
        best_lx, best_ly, best_hit, best_sec_dy = None, None, 0, float("inf")
        for wi in word_list:
            wn = wi["norm"]
            if sig[0] not in wn and wn not in sig[0]:
                continue
            if len(wn) < 3:
                continue
            hit = len(wn)
            sec_dy = 0.0  # mean y-distance of matched secondary words to anchor
            if len(sig) > 1:
                sec_matches = []
                for wj in word_list:
                    if abs(wj["cy"] - wi["cy"]) > SAME_ROW:
                        continue
                    if wj["x_min"] <= wi["x_max"]:
                        continue
                    if any(sw in wj["norm"] for sw in sig[1:] if len(sw) >= 3):
                        hit += len(wj["norm"])
                        sec_matches.append(abs(wj["cy"] - wi["cy"]))
                if sec_matches:
                    sec_dy = sum(sec_matches) / len(sec_matches)
            # prefer higher hit count; break ties by picking the anchor whose
            # secondary words are closest vertically (most same-row)
            if hit > best_hit or (hit == best_hit and sec_dy < best_sec_dy):
                best_hit    = hit
                best_lx     = wi["x_min"]
                best_ly     = wi["cy"]
                best_sec_dy = sec_dy
        if best_lx is not None:
            positions[label] = (best_lx, best_ly)

    return positions


def _find_compartment_via_pixel_detection(page_pdf_bytes: bytes, page) -> tuple[str, str]:
    """Returns (compartment, category). category is 'Móvel'|'Industrial'|'' (score>=150)."""
    import fitz
    import numpy as np
    from PIL import Image
    import io
    import cv2

    if not page_pdf_bytes or page is None:
        return "", ""

    try:
        doc      = fitz.open(stream=page_pdf_bytes, filetype="pdf")
        pix      = doc[0].get_pixmap(dpi=200, alpha=False)
        doc.close()
        png_bytes = pix.tobytes("png")
        img      = Image.open(io.BytesIO(png_bytes)).convert("L")
        img_arr  = np.array(img, dtype=np.float32)
        gray_u8  = np.array(img, dtype=np.uint8)
        img_rgb  = np.array(Image.open(io.BytesIO(png_bytes)).convert("RGB"), dtype=np.float32)
        h_px, w_px = img_arr.shape

        ph = getattr(page, "height", 11.0) or 11.0
        pw = getattr(page, "width",  8.5)  or 8.5
        scale_x = w_px / pw
        scale_y = h_px / ph

        COMP_X_MIN = pw * 0.55
        COMP_Y_MIN = ph * 0.06
        COMP_Y_MAX = ph * 0.94

        word_list = []
        for w in (getattr(page, "words", None) or []):
            poly = w.polygon
            if not poly:
                continue
            xs = [p.x for p in poly]
            ys = [p.y for p in poly]
            cx = sum(xs) / len(xs)
            cy = sum(ys) / len(ys)
            if cx < COMP_X_MIN or cy < COMP_Y_MIN or cy > COMP_Y_MAX:
                continue
            word_list.append({
                "norm": _norm(w.content),
                "cx": cx, "cy": cy,
                "x_min": min(xs), "x_max": max(xs),
            })

        marks = []
        for sm in (getattr(page, "selection_marks", None) or []):
            poly = sm.polygon
            if not poly:
                continue
            xs  = [p.x for p in poly]
            ys  = [p.y for p in poly]
            scx = sum(xs) / len(xs)
            scy = sum(ys) / len(ys)
            if scx < COMP_X_MIN or scy < COMP_Y_MIN or scy > COMP_Y_MAX:
                continue
            marks.append((scx, scy))

        label_pos = _find_label_positions(word_list)

        # Category (Móvel/Industrial) via ink-color scan — computed once, independent
        # of which phase finds the specific compartment.
        cat = _category_from_ink(label_pos, img_rgb, scale_x, scale_y, w_px, h_px, COMP_X_MIN)

        # ── PHASE 1: Azure marks ─────────────────────────────────────────
        circle_positions_p1 = []
        if marks and label_pos:
            Y_MARK = 0.20
            for label, (lx, ly) in label_pos.items():
                best_mark, best_ydist = None, float("inf")
                for scx, scy in marks:
                    if scx >= lx - 0.05:
                        continue
                    ydist = abs(scy - ly)
                    if ydist > Y_MARK or ydist >= best_ydist:
                        continue
                    best_ydist = ydist
                    best_mark  = (scx, scy)
                if best_mark:
                    circle_positions_p1.append((best_mark[0], best_mark[1], label))

        circle_data_p1 = _measure_circle_data(circle_positions_p1, img_arr, scale_x, scale_y, w_px, h_px)

        if circle_data_p1:
            top_p1  = sorted({l: round(s, 1) for s, l in circle_data_p1}.items(), key=lambda x: -x[1])[:5]
            sorted_p1 = sorted(circle_data_p1, key=lambda x: -x[0])
            gap_p1  = round(sorted_p1[0][0] - (sorted_p1[1][0] if len(sorted_p1) > 1 else 0), 1)
            _p1_msg = "P1-Azure: " + " | ".join(f"{l}:{s}" for l, s in top_p1) + f" | gap={gap_p1}"
            if _PM_DEBUG:
                print(f"  [P1_DBG] {_p1_msg}")
                # Also dump label_pos for eixo/movel
                _lp_dbg = {k: v for k, v in label_pos.items() if any(t in k.lower() for t in ("ixo", "vel", "raxa", "otor", "vel"))}
                print(f"  [P1_DBG] label_pos: {_lp_dbg}")
                print(f"  [P1_DBG] marks: {marks}")

        result = _pick_winner(circle_data_p1)
        if result:
            return result, cat
        # Do NOT bail out here when Phase 1 tied — Azure marks can sit 0.05-0.10 in
        # below the true circle centre, causing off-by-one label matches.  Phase 2
        # Hough finds the actual circle centre from pixels and resolves the tie.

        if not label_pos:
            return "", cat

        cx_min_px = int(COMP_X_MIN * scale_x)
        cy_min_px = int(COMP_Y_MIN * scale_y)
        cy_max_px = int(COMP_Y_MAX * scale_y)

        clahe    = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray_u8)

        panel_enhanced = enhanced[cy_min_px:cy_max_px, cx_min_px:]
        edges          = cv2.Canny(panel_enhanced, 40, 120)
        edges_dilated  = cv2.dilate(edges, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2, 2)))

        min_r = max(4, int(0.07 * scale_x))
        max_r = int(0.22 * scale_x)

        raw = cv2.HoughCircles(
            edges_dilated, cv2.HOUGH_GRADIENT,
            dp=1, minDist=int(0.16 * scale_y),
            param1=200, param2=10,
            minRadius=min_r, maxRadius=max_r,
        )

        circle_positions_p2 = []
        Y_BAND = 0.16

        if raw is not None:
            detected = np.round(raw[0]).astype(int)
            for cx_panel, cy_panel, _ in detected:
                cx_page = (cx_panel + cx_min_px) / scale_x
                cy_page = (cy_panel + cy_min_px) / scale_y
                best_label, best_ydist = None, float("inf")
                for lbl, (lx, ly) in label_pos.items():
                    if lx <= cx_page or lx - cx_page > 2.5:
                        continue
                    ydist = abs(ly - cy_page)
                    if ydist <= Y_BAND and ydist < best_ydist:
                        best_ydist = ydist
                        best_label = lbl
                if best_label:
                    circle_positions_p2.append((cx_page, cy_page, best_label))

        if not circle_positions_p2:
            # Offset fallback: only run when Azure detected at least one selection mark
            # in the compartment area — that's our signal that something was actually marked.
            # Without any marks, scanning all label offsets causes false positives from
            # dark form elements (borders, barcodes) near label positions.
            if not marks:
                return "", cat
            OFFSETS = [0.22, 0.28, 0.34, 0.40]
            for lbl, (lx, ly) in label_pos.items():
                best_off_x, best_off_score = lx - 0.28, -1.0
                for off in OFFSETS:
                    cx_try = lx - off
                    if cx_try < COMP_X_MIN:
                        continue
                    xi1 = max(0, int((cx_try - 0.06) * scale_x))
                    xi2 = min(w_px, int((cx_try + 0.06) * scale_x))
                    yi1 = max(0, int((ly - 0.06) * scale_y))
                    yi2 = min(h_px, int((ly + 0.06) * scale_y))
                    if xi1 >= xi2 or yi1 >= yi2:
                        continue
                    patch = img_arr[yi1:yi2, xi1:xi2]
                    dark  = float(255.0 - np.mean(patch))
                    if dark > best_off_score:
                        best_off_score = dark
                        best_off_x     = cx_try
                circle_positions_p2.append((best_off_x, ly, lbl))

        if not circle_positions_p2:
            return "", cat

        circle_data_p2 = _measure_circle_data(circle_positions_p2, img_arr, scale_x, scale_y, w_px, h_px)

        if not circle_data_p2:
            return "", cat
        sorted_p2  = sorted(circle_data_p2, key=lambda x: -x[0])
        scores_p2  = [s for s, _ in circle_data_p2]
        median_p2  = sorted(scores_p2)[len(scores_p2) // 2]
        best_p2, label_p2 = sorted_p2[0]
        second_p2 = sorted_p2[1][0] if len(sorted_p2) > 1 else 0.0
        if best_p2 < 155 or (best_p2 - median_p2) < 65 or (best_p2 - second_p2) < 55:
            return "", cat
        return COMPARTMENT_OPTIONS.get(_norm(label_p2), label_p2), cat

    except Exception:
        return "", ""


SYNONYMS = {
    # Chassi
    "seriechassi":        "Chassi Série",
    "serie/chassi":       "Chassi Série",
    "serie chassi":       "Chassi Série",
    "serie do chassi":    "Chassi Série",
    "chassi serie":       "Chassi Série",
    "chasis":             "Chassi Série",
    "serie/chasis":       "Chassi Série",

    # Tag frota
    "frota/tag":  "Tag Frota",
    "frota tag":  "Tag Frota",
    "tag frota":  "Tag Frota",
    "tag":        "Tag Frota",
    "equipo":     "Tag Frota",
    "equipo tag": "Tag Frota",
    "flota":      "Tag Frota",
    "flota/tag":  "Tag Frota",

    # Ponto coleta
    "ponto de coleta":               "Ponto de Coleta / Compartimento",
    "ponto de coleta compartimento": "Ponto de Coleta / Compartimento",
    "ponto coleta":                  "Ponto de Coleta / Compartimento",
    "compartimento":                 "Ponto de Coleta / Compartimento",
    "tipo de compartimento":         "Ponto de Coleta / Compartimento",
    "tipo compartimento":            "Ponto de Coleta / Compartimento",

    # Horimetro
    "horimetro/km/periodo": "Horímetro/Km/Período",
    "horimetro km periodo": "Horímetro/Km/Período",
    "horimetro":            "Horímetro/Km/Período",
    "horometro":            "Horímetro/Km/Período",
    "km/periodo":           "Horímetro/Km/Período",

    # Óleo trocado
    "oleo trocado":    "Óleo trocado",
    "óleo trocado":    "Óleo trocado",
    "fluido trocado":        "Óleo trocado",
    "fluído trocado":        "Óleo trocado",
    "fluido trocado?":       "Óleo trocado",
    "fluído trocado?":       "Óleo trocado",
    "fluido trocado? sim":   "Óleo trocado",
    "fluido trocado? nao":   "Óleo trocado",
    "fluido trocado? não":   "Óleo trocado",
    "fluido trocado? si":    "Óleo trocado",
    "fluido trocado? no":    "Óleo trocado",
    "aceite cambiado":  "Óleo trocado",
    "aceite cambiado?": "Óleo trocado",

    # Número do frasco
    "amostra": "Número do Frasco",
    "muestra": "Número do Frasco",

    # Código externo
    "codigo ext/os":   "Código externo",
    "codigo ext./os":  "Código externo",
    "codigo ext os":   "Código externo",
    "codigo ext/ot":   "Código externo",
    "codigo ext./ot":  "Código externo",
    "codigo ext ot":   "Código externo",
    "codigo externo":  "Código externo",

    # Data
    "data da coleta":      "Data da Coleta",
    "data coleta":         "Data da Coleta",
    "fecha de muestreo":   "Data da Coleta",
    "fecha muestreo":      "Data da Coleta",
    "fecha de muestra":    "Data da Coleta",
    "fecha":               "Data da Coleta",

    # Volume adicionado
    "vol oleo adic":          "Volume adicionado",
    "vol. oleo adic":         "Volume adicionado",
    "vol fluido adic":        "Volume adicionado",
    "vol. fluido adic":       "Volume adicionado",
    "vol fluido adicionado":  "Volume adicionado",
    "vol. fluido adicionado": "Volume adicionado",

    # Viscosidade
    "viscosidade":       "Viscosidade (Óleo)",
    "viscosidade oleo":  "Viscosidade (Óleo)",
    "viscosidad":        "Viscosidade (Óleo)",

    # Descrição / Fabricante+Modelo do óleo → always into Fabricante field
    # so the consolidation step can assemble them cleanly
    "fabricante e modelo":    "Fabricante (Óleo)",
    "fabricante y modelo":    "Fabricante (Óleo)",
    "descripcion del aceite": "Fabricante (Óleo)",
    "descricao do oleo":      "Fabricante (Óleo)",
    "descricao do óleo":      "Fabricante (Óleo)",
    "descricao":              "Fabricante (Óleo)",

    # Horas/km do fluído
    "horas/km do fluido":   "Horas/Km do Fluído",
    "horas km do fluido":   "Horas/Km do Fluído",
    "horas/km de aceite":   "Horas/Km do Fluído",
    "horas km de aceite":   "Horas/Km do Fluído",
    "horas/km do oleo":     "Horas/Km do Fluído",
    "horas km do oleo":     "Horas/Km do Fluído",
    "hrs/km do fluido":     "Horas/Km do Fluído",
    "hrs km do fluido":     "Horas/Km do Fluído",
    "hrs/km fluido":        "Horas/Km do Fluído",
    "hrs km fluido":        "Horas/Km do Fluído",
    "horas do fluido":      "Horas/Km do Fluído",
    "horas fluido":         "Horas/Km do Fluído",
    "km do fluido":         "Horas/Km do Fluído",
    "hrs fluido":           "Horas/Km do Fluído",
    "hrs do fluido":        "Horas/Km do Fluído",
    "horas/km fluido":      "Horas/Km do Fluído",
    "horas km fluido":      "Horas/Km do Fluído",
    "hrs/km oleo":          "Horas/Km do Fluído",
    "hrs km oleo":          "Horas/Km do Fluído",

    # Comentário
    "observacoes/feedback": "Comentário",
    "observacoes":          "Comentário",
    "observaciones":        "Comentário",
    "comentario":           "Comentário",
    "comentarios":          "Comentário",

    # Compartment variants
    "reducao final esq":       "Redução final Esq.",
    "reducao final dir":       "Redução final Dir.",
    "reducao final esq.":      "Redução final Esq.",
    "reducao final dir.":      "Redução final Dir.",
    "reducao final esquerd":   "Redução final Esq.",
    "reducao final direit":    "Redução final Dir.",
    "comando final dir":       "Comando Final Dir.",
    "comando final esq":       "Comando Final Esq.",
    "comando final dir.":      "Comando Final Dir.",
    "comando final esq.":      "Comando Final Esq.",
    "direcao":                 "Direção",
    "sistema de lubrif":       "Sistema de Lubr.",
    "sistema de lubrif.":      "Sistema de Lubr.",
    "sistema lubr":            "Sistema de Lubr.",
    "graxa industrial":        "Graxa Industrial",
    "compressor industrial":   "Compressor Industrial",
    "hidraulico industrial":   "Hidráulico Industrial",
    "hidroulico industrial":   "Hidráulico Industrial",
    "eixo traseiro":           "Eixo Traseiro",
    "eixo dianteiro":          "Eixo Dianteiro",
    "ponto de coleta:":        "Ponto de Coleta",
    "ponto coleta:":           "Ponto de Coleta",

    # Tanque — OCR misreadings
    "tanque":  "Tanque",
    "tonque":  "Tanque",
    "tonopo":  "Tanque",
    "trenque": "Tanque",
    "granque": "Tanque",
    "janque":  "Tanque",
    "tanqe":   "Tanque",
    "tanqu":   "Tanque",
    "tanquo":  "Tanque",
    "tanque.": "Tanque",

    # Fabricante / Modelo — only oil-qualified names (never bare "fabricante" or "modelo")
    "fabricante oleo":        "Fabricante (Óleo)",
    "fabricante (oleo)":      "Fabricante (Óleo)",
    "fabricante/modelo":      "Fabricante (Óleo)",
    "fabricante / modelo":    "Fabricante (Óleo)",
    "fabricante modelo":      "Fabricante (Óleo)",
    "fabricante e modelo oleo": "Fabricante (Óleo)",
    "modelo oleo":            "Fabricante (Óleo)",   # "modelo oleo" = oil model → same field
    "modelo (oleo)":          "Fabricante (Óleo)",
    # NOTE: bare "fabricante" and bare "modelo" are intentionally NOT mapped here —
    # those are equipment fields (above the FLUIDO line) and must not pollute oil info.
}


def _empty_if_none_like(v: str) -> str:
    vv = (v or "").strip()
    if vv.lower() in ("none", "null", "nan", "-", "n/a", "unselected", "unreadable", "illegible"):
        return ""
    return vv


def _expand_two_digit_year(yy: str) -> str:
    yy = yy.strip()
    if len(yy) == 2 and yy.isdigit():
        return "20" + yy
    return yy


def _normalize_date_str(v: str) -> str:
    v = _empty_if_none_like(v)
    if not v:
        return ""
    vv = re.sub(r"[a-zA-ZÀ-ÿ]", "", v).strip()
    if not vv:
        return ""
    m = re.search(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", vv)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        y = _expand_two_digit_year(y)
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    digits = re.sub(r"\D", "", vv)
    if len(digits) == 8:
        d, mo, y = digits[:2], digits[2:4], digits[4:]
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    if len(digits) == 6:
        d, mo, y = digits[:2], digits[2:4], "20" + digits[4:]
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    return ""


_OCR_DIGIT_MAP = {
    ord("O"): "0", ord("o"): "0",
    ord("I"): "1", ord("i"): "1", ord("L"): "1", ord("l"): "1",
    ord("S"): "5", ord("s"): "5",
    ord("B"): "8",
    ord("Z"): "2", ord("z"): "2",
    ord("G"): "6",
    ord("g"): "9",
    ord("T"): "7",
    ord("Q"): "0", ord("q"): "9",
}


def _clean_numeric(v: str) -> str:
    v = v.translate(_OCR_DIGIT_MAP)
    return re.sub(r"[^\d]", "", v)


def _clean_horimetro(v: str) -> str:
    """
    Clean a horimeter/km value, preserving the unit suffix when present.

    Examples:
      "160262h"    → "160262h"
      "160262 H"   → "160262h"
      "160.262h"   → "160262h"   (period = thousands separator)
      "160,262h"   → "160262h"   (comma  = thousands separator)
      "160262/h"   → "160262h"   (slash before unit)
      "160/262"    → "160262"    (slash as thousands separator)
      "1.602,62h"  → "160262h"   (European format)
      "45000 KM"   → "45000km"
      "1602G2"     → "160262"    (OCR digit fix)
    """
    v = v.strip()
    if not v:
        return ""

    # Detect units from ORIGINAL string (before _OCR_DIGIT_MAP which turns
    # 'L'→'1', 'S'→'5', 'H' → left as-is but 'S' would corrupt "HS"→"H5").
    v_orig_up = v.upper()

    unit = ""
    v_numeric = v   # original string with the unit word stripped out

    if re.search(r"\b(KMS?)\b", v_orig_up) or v_orig_up.rstrip(" ./").endswith("KM"):
        unit = "km"
        v_numeric = re.sub(r"\b(KMS?)\b", "", v, flags=re.I).strip()
    elif re.search(r"\b(HORAS?)\b", v_orig_up):
        unit = "h"
        v_numeric = re.sub(r"\b(HORAS?)\b", "", v, flags=re.I).strip()
    elif re.search(r"\d\s*(HRS?|HS)\s*$", v_orig_up):
        # "500HRS", "500HS", "500 HS" — digit + H suffix with no word boundary before H
        unit = "h"
        v_numeric = re.sub(r"(HRS?|HS)\s*$", "", v, flags=re.I).strip()
    elif re.search(r"\b(HRS?)\b", v_orig_up):
        # "500 HR" / "500 HRS" — space before, so word boundary works
        unit = "h"
        v_numeric = re.sub(r"\b(HRS?)\b", "", v, flags=re.I).strip()
    elif re.search(r"\d\s*[/\s]*H\s*$", v_orig_up):
        # "500H", "500 H", "500/H" — bare H at end
        unit = "h"
        v_numeric = re.sub(r"H\s*$", "", v, flags=re.I).strip()
    elif re.search(r"\b(ML|MILILITROS?)\b", v_orig_up):
        unit = "ml"
        v_numeric = re.sub(r"\b(ML|MILILITROS?)\b", "", v, flags=re.I).strip()
    elif re.search(r"\b(LTS?|LITROS?|LITOS?|LITO)\b", v_orig_up):
        unit = "L"
        v_numeric = re.sub(r"\b(LTS?|LITROS?|LITOS?|LITO)\b", "", v, flags=re.I).strip()
    elif re.search(r"\d\s*L$", v_orig_up):
        unit = "L"
        v_numeric = re.sub(r"L\s*$", "", v, flags=re.I).strip()
    elif re.match(r"^L\s*\d", v_orig_up):
        # "L04", "L 04", "L8" — L as a prefix (instead of suffix) abbreviating Litros
        unit = "L"
        v_numeric = re.sub(r"^L\s*", "", v, flags=re.I).strip()

    # Apply OCR digit substitutions to the NUMERIC part only
    v_fixed = v_numeric.translate(_OCR_DIGIT_MAP)

    # ── Normalize thousands separators ───────────────────────────────────
    # Pattern: digit(s) + separator(.,/) + exactly 3 digits → remove separator
    # Handles: 160.262 / 160,262 / 160/262 / 1.602,62 (European)
    # Repeat to handle multiple groups: 1.602.262 → 1602262
    for _ in range(4):
        v_fixed = re.sub(r"(\d)[.,/](\d{3})(?=\D|$)", r"\1\2", v_fixed)

    # ── Extract digits only ───────────────────────────────────────────────
    digits = re.sub(r"[^\d]", "", v_fixed)
    if not digits:
        return ""

    return digits + unit


_VOLUME_TEXT_OK = {
    "cheio", "completo", "nao adicionado", "sem adicao",
    "nenhum", "zero", "nao", "sem", "vazio",
}

# Visual-confusion character classes for OCR misreads of handwritten "Lts"
# (abbreviation of Litros): each position can be read as any visually similar shape.
_LTS_L_LIKE = set("l1it")
_LTS_T_LIKE = set("tfei")
_LTS_S_LIKE = set("s5")


def _looks_like_lts_abbrev(suffix: str) -> bool:
    """True if a 3-letter suffix is a plausible OCR misread of 'Lts'."""
    s = suffix.lower()
    if len(s) != 3:
        return False
    return s[0] in _LTS_L_LIKE and s[1] in _LTS_T_LIKE and s[2] in _LTS_S_LIKE

def _clean_volume(v: str) -> str:
    """
    Clean 'Volume adicionado'. Accepts:
      - Numbers with optional unit:  '80L', '1.500 LTS', '500 ml', '2750'
      - Alphanumeric codes:          'CT50', 'CTS0'
      - Known text values:           'cheio', 'completo', 'não adicionado'
      - Oil-grade references:        '15W90', '85W140'  (SAE grade in volume field)
      - Qualified amounts:           '500 HS', '500 DRS' (amount + qualifier)
    Rejects:
      - Form label leakage:          'adic.', 'Jido adic. 50h', 'vol fluido'
      - Operator noise:              '(+90', 'C+50'
      - Pure text with no digits:    'etsa'
    """
    v = v.strip()
    if not v:
        return ""

    # Known text values → keep as-is
    if _norm(v) in _VOLUME_TEXT_OK:
        return v

    # Reject if contains form-label words
    vn = _norm(v)
    for label_word in ("adic", "fluido", "oleo", "vol ", "jido", "lido"):
        if label_word in vn:
            return ""

    # Reject if contains operator/parenthesis noise
    if re.search(r"[+\(\)<>]", v):
        return ""

    # Reject if no digits at all (pure text like "etsa")
    digits = re.sub(r"[^\d]", "", v)
    if not digits:
        return ""

    # SAE viscosity grades (e.g. "15W90", "85W140") → keep as oil-grade reference
    if re.search(r"\d+[Ww]\d+", v):
        return v

    # Try to parse OCR-corrupted SAE grades: "15uga" → "15W40", "15mno" → "15W40"
    # Collapse spaces before trying so "15 uga" → "15uga"
    _v_collapsed = re.sub(r"\s+", "", v)
    _vis = _parse_viscosity(_v_collapsed)
    if _vis:
        return _vis

    # Volume units detected early → always normalize via _clean_horimetro
    # so "2 Litros", "3 LTS", "500 ml" work even when digit ratio is low.
    # Also catches "L04"/"L 04" (L as a prefix abbreviating Litros).
    if re.search(r"\b(LITROS?|LTS?|LITO[S]?|ML|MILILITROS?)\b", v.upper()) or re.match(r"^L\s*\d", v.upper()):
        return _clean_horimetro(v)

    # Garbled handwritten "Lts" abbreviation: OCR reads the handwritten L/T/S shapes
    # as other similarly-shaped letters → "04 lis", "08 tes", "13 tts" are all
    # attempts at "Lts". Match by visual-confusion character classes instead of an
    # exact word list.
    _m_lts = re.match(r"^(\d+)\s*([A-Za-z]{3})$", v)
    if _m_lts and _looks_like_lts_abbrev(_m_lts.group(2)):
        return _clean_horimetro(_m_lts.group(1) + " Lts")

    # Primarily numeric (>70% digits) → full cleaning + unit detection.
    # Threshold raised from 0.50 to 0.70 so qualified amounts like "500 HS" or
    # "500 DRS" are kept as-is instead of being stripped to bare digits.
    if len(digits) / max(len(v), 1) > 0.70:
        return _clean_horimetro(v)

    # Alphanumeric code (e.g. "CT50", "500 HS", "500 DRS") → keep as-is
    return v


def clean_value(col_name: str, value: str) -> str:
    v = _empty_if_none_like(value)
    if not v:
        return ""

    if col_name == "Tag Frota":
        return re.sub(r"[=–—]", "-", v)

    if col_name == "Data da Coleta":
        return _normalize_date_str(v)

    if col_name in ("Número do Frasco", "Código externo"):
        m = re.search(r"\b(\d{6,})\b", v)
        return m.group(1) if m else v

    if col_name == "Óleo trocado":
        vn = _norm(v)
        if vn in ("sim", "si", "yes", "selected", "true", "1", "x"):
            return "Sim"
        if vn in ("nao", "não", "no"):
            return "Não"

    if col_name in ("Horímetro/Km/Período", "Horas/Km do Fluído"):
        cleaned = _clean_horimetro(v)
        if col_name == "Horas/Km do Fluído":
            # Volume units (L, ml) are never valid for a horimeter/km field
            if cleaned.endswith(("L", "ml")):
                return ""
            # Volumes are 1-2 digits; horimeter readings are always ≥3 digits
            if cleaned and len(re.sub(r"\D", "", cleaned)) < 3:
                return ""
        return cleaned

    if col_name == "Volume adicionado":
        return _clean_volume(v)

    return v


def key_to_column(k: str) -> str | None:
    kn = _norm(k)
    if kn in SYNONYMS:
        return SYNONYMS[kn]
    if kn in COL_NORM_MAP:
        return COL_NORM_MAP[kn]
    best = get_close_matches(kn, COL_NORMS, n=1, cutoff=0.78)
    if best:
        col = COL_NORM_MAP[best[0]]
        # Guard: generic "fabricante" or "modelo" (equipment fields) must NOT be
        # fuzzy-matched into oil columns. Only names that contain an oil qualifier
        # (e.g. "fabricante e modelo", "fabricante oleo", "viscosidade") should map there.
        if col in ("Fabricante (Óleo)", "Modelo (Óleo)", "Viscosidade (Óleo)"):
            oil_hint = any(x in kn for x in (
                "oleo", "fluido", "visc", "e modelo", "y modelo", "/modelo", "lubr",
            ))
            if not oil_hint:
                return None
        return col
    return None


def _field_plausibility_cap(col_name: str, value: str) -> float:
    """Rule-based plausibility ceiling (0-100) for a cell, INDEPENDENT of the form
    layout — it only checks whether the extracted value has the shape expected for
    that field type. Returns a CAP: 100 = no objection; lower = the value looks
    implausible and the cell's reliability should not exceed this. Combined with the
    Azure/pixel confidence via min() so a confident-but-implausible read still gets
    flagged for review (e.g. viscosity "154140", date "99/99/99")."""
    v = (value or "").strip()
    if not v:
        return 100.0

    if col_name == "Categoria":
        return 100.0 if v in CATEGORY_LABELS else 40.0

    if col_name == "Óleo trocado":
        return 100.0 if v in ("Sim", "Não") else 40.0

    if col_name == "Ponto de Coleta / Compartimento":
        # Plausible when it's a known compartment label.
        return 100.0 if _norm(v) in COMPARTMENT_OPTIONS else 60.0

    if col_name == "Número do Frasco":
        digits = re.sub(r"\D", "", v)
        if len(digits) >= 8:   # frascos are ~10 digits
            return 100.0
        if len(digits) >= 5:
            return 65.0
        return 35.0

    if col_name in ("Horímetro/Km/Período", "Horas/Km do Fluído"):
        digits = re.sub(r"\D", "", v)
        if not digits:
            return 30.0
        compact = re.sub(r"\s", "", v)
        ratio = len(digits) / max(len(compact), 1)
        return 100.0 if ratio >= 0.6 else 55.0

    if col_name == "Volume adicionado":
        if _norm(v) in _VOLUME_TEXT_OK:
            return 100.0
        if re.match(r"^\d+([.,]\d+)?\s*(L|ml|litros?)?$", v, re.I):
            return 100.0
        return 70.0 if re.search(r"\d", v) else 35.0

    if col_name == "Data da Coleta":
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", v)
        if not m:
            return 45.0
        d, mo = int(m.group(1)), int(m.group(2))
        return 100.0 if (1 <= d <= 31 and 1 <= mo <= 12) else 40.0

    if col_name == "Fabricante/Modelo":
        # A bare 4+ digit run is a mis-read viscosity that leaked into the name.
        for tok in v.split():
            if re.match(r"^\d{4,}$", tok):
                return 50.0
        return 100.0

    if col_name == "Viscosidade (Óleo)":
        # Plausible viscosity shapes: SAE (15W40 / 80W90 / SAE 30), ISO VG nnn,
        # or a plain grade number. A long bare digit run (e.g. "154140") is junk.
        vu = v.upper().strip()
        if re.search(r"\d{1,3}\s*W\s*\d{0,3}", vu):       # 15W40, 80W90, 20W
            return 100.0
        if re.search(r"\bISO\b|\bVG\b|\bSAE\b|\bNLGI\b|\bAW\b|\bEP\b|\bHD\b", vu):
            return 100.0
        if re.match(r"^\d{2,4}$", vu):                     # 220, 460, 68
            return 100.0
        if re.search(r"\d{5,}", vu):                       # 154140 → junk
            return 45.0
        return 70.0

    # Chassi, Tag Frota, Comentário, Código externo: free text → no shape rule.
    return 100.0


def fields_to_record(fields: dict, page=None, confidences: dict = None, page_bytes: bytes = None, page_mime: str = "application/pdf") -> dict:
    record = {c: "" for c in COLETA_COLUMNS}

    # Tracks confidence (0-100) per COLETA_COLUMN, regardless of which detection
    # path produced the value (Azure KV, pixel/ink detection, or text fallback).
    # Stashed into record["__confidence__"] at the end for the "Revisar" column.
    _conf: dict[str, float] = {}

    def is_selected(v) -> bool:
        return str(v).strip().lower() in ("selected", "sim", "si", "yes", "true", "1", "x")

    # --- First pass: compartment selection marks (Azure KV) ---
    # Collect ALL selected compartments so the Móvel/Industrial filter works:
    # when Móvel or Industrial is co-selected with a specific compartment, return
    # the specific one; only return Móvel/Industrial when nothing else is selected.
    _kv_hits: list[tuple[str, float]] = []
    for k, v in (fields or {}).items():
        if v is None:
            continue
        kn        = _norm(k)
        canonical = COMPARTMENT_OPTIONS.get(kn)
        if not canonical:
            continue
        conf = (confidences or {}).get(k, 1.0) or 1.0
        if is_selected(v) and conf >= 0.70:
            _kv_hits.append((canonical, conf))

    if _kv_hits:
        _kv_hits.sort(key=lambda x: -x[1])          # highest confidence first
        _kv_specific = [c for c, _ in _kv_hits if c not in CATEGORY_LABELS]
        record["Ponto de Coleta / Compartimento"] = (
            _kv_specific[0] if _kv_specific else _kv_hits[0][0]
        )
        _conf["Ponto de Coleta / Compartimento"] = _kv_hits[0][1] * 100

    # --- Pixel verification of KV compartment ---
    # Azure's custom KV model sometimes returns "selected" for checkboxes that are
    # visually empty.  Use page_marks (which verifies pixel darkness at the actual
    # selection-mark position) to confirm the KV result.  If page_marks finds
    # nothing with pixel verification, the KV hit is likely a false positive.
    if record["Ponto de Coleta / Compartimento"] and page_bytes is not None and page is not None:
        pm_verify = _find_compartment_via_page_marks(page, page_bytes=page_bytes)
        # Always trust the pixel-verified result: if pm_verify found something use it;
        # if nothing found, the KV hit was a false positive → clear it.
        record["Ponto de Coleta / Compartimento"] = pm_verify
        if pm_verify:
            _conf["Ponto de Coleta / Compartimento"] = max(_conf.get("Ponto de Coleta / Compartimento", 0.0), 85.0)
        else:
            _conf.pop("Ponto de Coleta / Compartimento", None)

    # --- Pre-pass: Óleo trocado selection marks ---
    # Must run BEFORE the empty-value filter because Azure often returns empty
    # for these checkboxes even when one is physically checked. When the key
    # itself encodes "Sim" or "Não" (e.g. "Fluido Trocado? Sim"), we can
    # resolve the answer from the key name + selected state.
    for k, v in (fields or {}).items():
        kn = _norm(k)
        is_oleo_key = ("oleo trocado" in kn) or ("fluido trocado" in kn) or ("aceite cambiado" in kn)
        if not is_oleo_key:
            continue
        v_str   = str(v or "").strip()
        v_clean = _norm(v_str)
        _oleo_conf_pct = ((confidences or {}).get(k, 0.90) or 0.90) * 100

        # Explicit value carries the answer
        if v_clean in ("sim", "si", "yes", "selected", "true", "1", "x"):
            # Don't assign "Sim" if the key explicitly says "Não"
            if not (kn.endswith(" nao") or kn.endswith(" no") or "? nao" in kn or "? no" in kn):
                record["Óleo trocado"] = "Sim"
                _conf["Óleo trocado"] = _oleo_conf_pct
                break
        if v_clean in ("nao", "não", "no"):
            if not (kn.endswith(" sim") or kn.endswith(" si") or "? sim" in kn or "? si" in kn):
                record["Óleo trocado"] = "Não"
                _conf["Óleo trocado"] = _oleo_conf_pct
                break
        if "sim" in v_clean and "nao" not in v_clean:
            record["Óleo trocado"] = "Sim"
            _conf["Óleo trocado"] = _oleo_conf_pct
            break
        if "nao" in v_clean or ("no" in v_clean and "sim" not in v_clean):
            record["Óleo trocado"] = "Não"
            _conf["Óleo trocado"] = _oleo_conf_pct
            break

        # Key-encoded answer: "óleo trocado? Sim" / "Fluido Trocado? Não"
        # Azure returns empty but the key name tells us which checkbox this is.
        # Use is_selected to confirm it's checked.
        key_says_sim = kn.endswith(" sim") or kn.endswith(" si") or "? sim" in kn or "? si" in kn
        key_says_nao = kn.endswith(" nao") or kn.endswith(" no") or "? nao" in kn or "? no" in kn
        if key_says_sim and is_selected(v_str):
            record["Óleo trocado"] = "Sim"
            _conf["Óleo trocado"] = _oleo_conf_pct
            break
        if key_says_nao and is_selected(v_str):
            record["Óleo trocado"] = "Não"
            _conf["Óleo trocado"] = _oleo_conf_pct
            break

    # --- Main pass ---
    for k, v in (fields or {}).items():
        if v is None:
            continue
        v_str = str(v).strip()
        if _empty_if_none_like(v_str) == "":
            continue

        kn = _norm(k)

        if kn in COMPARTMENT_OPTIONS:
            continue

        if ("oleo trocado" in kn) or ("fluido trocado" in kn) or ("aceite cambiado" in kn):
            v_clean = _norm(v_str)
            _oleo_conf_pct2 = ((confidences or {}).get(k, 0.90) or 0.90) * 100
            if v_clean in ("sim", "si", "yes", "true", "1", "x"):
                record["Óleo trocado"] = "Sim"
                _conf["Óleo trocado"] = _oleo_conf_pct2
                continue
            if v_clean in ("nao", "não", "no"):
                record["Óleo trocado"] = "Não"
                _conf["Óleo trocado"] = _oleo_conf_pct2
                continue
            if is_selected(v_str):
                if (" sim" in kn) or kn.endswith(" sim") or (" si" in kn) or kn.endswith(" si"):
                    record["Óleo trocado"] = "Sim"
                    _conf["Óleo trocado"] = _oleo_conf_pct2
                    continue
                if (" nao" in kn) or (" no" in kn) or kn.endswith(" nao") or kn.endswith(" no"):
                    record["Óleo trocado"] = "Não"
                    _conf["Óleo trocado"] = _oleo_conf_pct2
                    continue
            if "sim" in v_clean:
                record["Óleo trocado"] = "Sim"
                _conf["Óleo trocado"] = _oleo_conf_pct2
                continue
            if "nao" in v_clean or "no" in v_clean:
                record["Óleo trocado"] = "Não"
                _conf["Óleo trocado"] = _oleo_conf_pct2
                continue
            continue

        col = key_to_column(k)
        if not col:
            continue

        cleaned = clean_value(col, v_str)
        _col_conf_pct = ((confidences or {}).get(k, 1.0) or 1.0) * 100

        record[col] = cleaned
        if cleaned:
            _conf[col] = _col_conf_pct

    # ── Clean / validate Chassi Série and Tag Frota ──────────────────────
    _FROTA_PREFIX = re.compile(
        r"^frota\s*/?\s*tag\s*[:\s(]?|^frota\s+tag\s*[:\s]?|^frotal?\s*[:/]?\s*",
        re.I,
    )
    _SERIE_PREFIX = re.compile(
        r"^s[eé]rie\s*/?\s*chassi\s*[:\s]?|^chassi\s*/?\s*s[eé]rie\s*[:\s]?|"
        r"^s[eé]rie\s*[:\s]?|^chassi\s*[:\s]?",
        re.I,
    )
    _COMP_LABEL = re.compile(
        r"^(eixo\s+(dianteiro|traseiro)|motor|transmissao|hidraulico|"
        r"comando\s+(final|de\s+giro)|compressor|diesel|direcao|"
        r"grax?a|redutor|turbina|bomba|mancal|isolante|tanque|"
        r"sistema\s+(de\s+lubr|termico)|radiador)$",
        re.I,
    )
    _LABEL_ONLY = re.compile(
        r"^(frota\s*[/\-]?\s*tags?|tags?\s*[/\-]?\s*frota|frotal?|tags?"
        r"|s[eé]rie\s*[/\-]?\s*chassi|chassi\s*[/\-]?\s*s[eé]rie"
        r"|s[eé]rie|chassi|chasis?|chassis"
        r"|equipamento|modelo|fabricante|cliente|cnpj|obra)[\s:/()]*$",
        re.I,
    )
    # Words that belong exclusively to the compartment section of the form.
    # If ANY of these appear as a token in Tag Frota / Chassi, the value is a
    # false positive (Azure read from the compartment panel instead of the ID field).
    _COMP_WORDS = frozenset({
        "eixo", "dianteiro", "traseiro", "traceiro",   # Eixo Dianteiro/Traseiro (+ OCR variant)
        "radiador",                                     # Radiador
        "transmissao",                                  # Transmissão
        "hidraulico",                                   # Hidráulico
        "compressor", "direcao", "diesel",              # more compartments
        "mancal", "turbina", "isolante",                # industrial
        "redutor", "reducao",                           # Redutor/Redução
        "comando",                                      # Comando Final
    })
    # OCR-mangled prefixes that still indicate a compartment word
    _COMP_PREFIXES = ("diant", "trasi", "traci", "dlant", "radio", "transm", "hidra", "manca")

    def _has_comp_word(text: str) -> bool:
        """True if text contains a compartment-section word → discard as false positive."""
        for w in _norm(text).split():
            if w in _COMP_WORDS:
                return True
            if len(w) >= 5 and any(w.startswith(p) for p in _COMP_PREFIXES):
                return True
        return False
    _CHASSI_PREFIX = re.compile(
        r"^s[eé]rie\s*[/\-]?\s*chassi[\s:/]*"
        r"|^chassi\s*[/\-]?\s*s[eé]rie[\s:/]*"
        r"|^s[eé]rie[\s:/]+"
        r"|^chassi[\s:/]+"
        r"|^frota\s*[/\-]?\s*tag[\s:/()]*"
        r"|^frota[\s:/]+"
        r"|^frotal?[\s:/]*",
        re.I,
    )

    def _clean_id_field(value: str, prefix_re, min_len: int = 2) -> str:
        v = value.strip()
        if not v:
            return ""
        stripped  = prefix_re.sub("", v).strip(" :/()").strip()
        if not stripped:
            candidate = "" if prefix_re.match(v) else v
        else:
            candidate = stripped
        if _LABEL_ONLY.match(candidate) or _COMP_LABEL.match(_norm(candidate)):
            return ""
        # Discard if any token is a compartment-section word (e.g. "Eixo Dianteiro Traseiro",
        # "Radiador cmicc", "Eixo Dianteiro Eixo Traseiro", etc.)
        if _has_comp_word(candidate):
            return ""
        if len(candidate) < min_len:
            return ""
        return candidate

    record["Chassi Série"] = _clean_id_field(record["Chassi Série"], _CHASSI_PREFIX, min_len=3)
    record["Tag Frota"]    = _clean_id_field(record["Tag Frota"],    _FROTA_PREFIX,  min_len=2)
    if not record["Chassi Série"]:
        _conf.pop("Chassi Série", None)
    if not record["Tag Frota"]:
        _conf.pop("Tag Frota", None)

    # Split merged Chassi + Tag Frota values.
    # When the chassis number is physically long on the form and "spills" into
    # the Tag Frota box, Azure sometimes reads both into a single Chassi field.
    # Example: "J8R00253 PC589" → Chassi="J8R00253", Tag Frota="PC589".
    if " " in record["Chassi Série"] and not record["Tag Frota"]:
        parts = record["Chassi Série"].split(None, 1)
        if len(parts) == 2 and len(parts[0]) >= 3:
            tag_candidate = _clean_id_field(parts[1].strip(), _FROTA_PREFIX, min_len=2)
            if tag_candidate:
                record["Chassi Série"] = parts[0]
                record["Tag Frota"]    = tag_candidate

    # Fallback: Chassi Série
    if not record["Chassi Série"] and page is not None:
        value = _fallback_read_field(page_bytes, page, ["chassi", "serie", "serial"])
        if value:
            value = _clean_id_field(clean_value("Chassi Série", value), _CHASSI_PREFIX, min_len=3)
            if value:
                record["Chassi Série"] = value
                _conf["Chassi Série"] = 70.0

    # Fallback: Tag Frota
    if not record["Tag Frota"] and page is not None:
        value = _fallback_read_field(page_bytes, page, ["frota", "tag"])
        if value:
            value = _clean_id_field(clean_value("Tag Frota", value), _FROTA_PREFIX)
            if value:
                record["Tag Frota"] = value
                _conf["Tag Frota"] = 70.0

    # Fallback: Horímetro/Km/Período
    if not record["Horímetro/Km/Período"] and page is not None:
        value = _fallback_read_field(page_bytes, page, [
            "horimetro", "horometro", "horim", "horom",
            "km/period", "km/per", "periodo", "horimetro/km",
        ])
        if value:
            cleaned = _clean_horimetro(value)
            if cleaned:
                record["Horímetro/Km/Período"] = cleaned
                _conf["Horímetro/Km/Período"] = 70.0

    # Fallback: Horas/Km do Fluído
    # This field has NO fallback elsewhere — Azure often misses it entirely.
    # Trigger words cover the common label abbreviations found in the forms.
    if not record["Horas/Km do Fluído"] and page is not None:
        # "horas" / "hrs" are specific to "HORAS/KM DO FLUIDO" — they do NOT
        # match "HORIMETRO" (starts "horim"), "FLUIDO TROCADO" or "VOL. FLUIDO".
        # search_below=True handles forms where the value is written under the label.
        value = _fallback_read_field(page_bytes, page, [
            "horas/km", "hrs/km", "horas", "hrs",
            "km/fluido", "km/oleo", "km/aceite",
        ], search_below=True)
        if value:
            cleaned = _clean_horimetro(value)
            # Reject volume units (L, ml) — Horas/Km is always h or km
            if cleaned and cleaned.endswith(("L", "ml")):
                cleaned = ""
            # Require ≥3 digits: oil-change volumes are 1-2 digits (1L–30L),
            # horimeter readings are always 3+ digits (250h, 5000km, etc.)
            if cleaned and len(re.sub(r"\D", "", cleaned)) < 3:
                cleaned = ""
            if cleaned:
                record["Horas/Km do Fluído"] = cleaned
                _conf["Horas/Km do Fluído"] = 70.0

    # Fallback: Volume adicionado
    if not record["Volume adicionado"] and page is not None:
        # Single-word triggers that actually appear in page.words.
        # Pass 3 of _read_field_from_words (compound-label crawl) will then
        # advance label_x_max past "fluido"/"oleo"/"adic." before scanning the value.
        value = _fallback_read_field(page_bytes, page, ["vol", "vol."])
        if value:
            cleaned = _clean_volume(value)
            if cleaned:
                record["Volume adicionado"] = cleaned
                _conf["Volume adicionado"] = 70.0

    # Fallback: pixel detection for compartment (most reliable for filled circles)
    if not record["Ponto de Coleta / Compartimento"] and page_bytes is not None and page is not None:
        comp, cat = _find_compartment_via_pixel_detection(page_bytes, page)
        if _PM_DEBUG and comp:
            print(f"  [PIXEL_DBG] pixel_detection returned comp={comp!r} cat={cat!r}")
        if comp:
            record["Ponto de Coleta / Compartimento"] = comp
            _conf["Ponto de Coleta / Compartimento"] = 90.0
        if cat:
            record["Categoria"] = cat
            _conf["Categoria"] = 90.0

    # Fallback: Azure selection-mark state (catches X marks that pixel misses)
    if not record["Ponto de Coleta / Compartimento"] and page is not None:
        comp = _find_compartment_via_page_marks(page, page_bytes=page_bytes)
        if comp:
            record["Ponto de Coleta / Compartimento"] = comp
            _conf["Ponto de Coleta / Compartimento"] = 80.0

    # Read text below "Ponto de Coleta" label
    if record["Ponto de Coleta / Compartimento"] in ("", "Ponto de Coleta") and page is not None:
        desc = _read_ponto_de_coleta_description(page)
        if _PM_DEBUG:
            print(f"  [DESC_DBG] description raw={desc!r} record_before={record['Ponto de Coleta / Compartimento']!r}")
        if desc:
            desc_norm = _norm(desc)
            if desc_norm in SYNONYMS:
                desc = SYNONYMS[desc_norm]
            elif desc_norm in COMPARTMENT_OPTIONS:
                desc = COMPARTMENT_OPTIONS[desc_norm]
            else:
                from difflib import get_close_matches as _gcm
                # Use a high cutoff to avoid false matches like "EXO INTERMEDIARIO"
                # fuzzy-matching to "Eixo Traseiro". Only accept near-exact matches.
                best = _gcm(desc_norm, list(COMPARTMENT_OPTIONS.keys()), n=1, cutoff=0.72)
                if best:
                    desc = COMPARTMENT_OPTIONS[best[0]]
                    if _PM_DEBUG:
                        print(f"  [DESC_DBG] fuzzy matched {desc_norm!r} -> {desc!r}")
                else:
                    # No compartment match. If "Ponto de Coleta" was already the
                    # selected option, the description IS the answer — use raw text.
                    # If we got here from a fallback path (empty record), don't guess.
                    if record["Ponto de Coleta / Compartimento"] != "Ponto de Coleta":
                        desc = ""
            if desc:
                record["Ponto de Coleta / Compartimento"] = desc
                _conf["Ponto de Coleta / Compartimento"] = 70.0

    if not record["Óleo trocado"] and page is not None:
        # Phase 1: pixel detection (label-first + inner/ring ratio score)
        oleo = _find_oleo_trocado_via_pixel_detection(page_bytes, page) if page_bytes else ""
        if oleo:
            _conf["Óleo trocado"] = 88.0
        # Phase 2: Azure mark state fallback (original logic)
        if not oleo:
            oleo = _find_oleo_trocado_via_page_marks(page)
            if oleo:
                _conf["Óleo trocado"] = 75.0
        if oleo:
            record["Óleo trocado"] = oleo

    # ── Correct Fabricante/Modelo against the oil reference catalog ──────
    # Fixes OCR typos via fuzzy match (e.g. "CASTROI" → "CASTROL") and flags
    # (lowers confidence on) values not found anywhere in the catalog.
    # Viscosidade is INTENTIONALLY excluded: it is too short (e.g. "15W40")
    # for safe fuzzy matching (it would map "18W40"→"85W40"), and the
    # _parse_viscosity / _fix_nonstandard_viscosity pipeline already corrects
    # it in a structured, format-aware way.
    _fab_map, _mod_map, _visc_map = _load_oil_reference()

    # The "Fabricante (Óleo)" field is fed by the form's "Fabricante e modelo"
    # field, so it frequently holds a MODEL name (e.g. "Rimula Rf4C") rather than
    # a manufacturer. Try the manufacturer catalog first, then fall back to the
    # model catalog (with trailing-viscosity stripping for cases like
    # "Rimula Rf4C 154140" → "RIMULA R4").
    _fab_val = record.get("Fabricante (Óleo)", "")
    if _fab_val:
        _corr, _matched = _fuzzy_match_reference(_fab_val, _fab_map, cutoff=0.82)
        if not _matched:
            _corr2, _matched2 = _fuzzy_match_reference(
                _fab_val, _mod_map, cutoff=0.82, strip_trailing_visc=True
            )
            if _matched2:
                _corr, _matched = _corr2, True
        if _matched:
            record["Fabricante (Óleo)"] = _corr
        else:
            _conf["Fabricante (Óleo)"] = min(_conf.get("Fabricante (Óleo)", 100.0), 55.0)

    _mod_val = record.get("Modelo (Óleo)", "")
    if _mod_val:
        _corr, _matched = _fuzzy_match_reference(
            _mod_val, _mod_map, cutoff=0.82, strip_trailing_visc=True
        )
        if _matched:
            record["Modelo (Óleo)"] = _corr
        else:
            _conf["Modelo (Óleo)"] = min(_conf.get("Modelo (Óleo)", 100.0), 55.0)

    # ── Split oil fields into TWO visible columns ────────────────────────
    # "Fabricante/Modelo" = Fabricante + Modelo (the name), normalized.
    # "Viscosidade (Óleo)" = the viscosity grade alone.
    # Reading each field independently (instead of merging into one) avoids the
    # OCR mixing the name with the viscosity number.

    def _fix_visc_spacing(s: str) -> str:
        """Fix spaced viscosity only: '15 W 40' → '15W40'. Nothing else."""
        s = re.sub(r"(?<![A-Za-z])(\d{1,3})\s*[Ww]\s*(\d{2,3})(?!\d)", r"\1W\2", s)
        return s.strip()

    # --- Fabricante/Modelo (name) ---
    _name_parts: list[str] = []
    for _p in (_fix_visc_spacing(record.get("Fabricante (Óleo)") or ""),
               _fix_visc_spacing(record.get("Modelo (Óleo)") or "")):
        if _p and _p.lower() not in " ".join(_name_parts).lower():
            _name_parts.append(_p)
    record["Fabricante/Modelo"] = " ".join(_name_parts)
    if record["Fabricante/Modelo"]:
        record["Fabricante/Modelo"] = _normalize_oil_description(record["Fabricante/Modelo"])

    # Confidence of the name = weakest of its parts.
    _name_confs = [
        _conf[c] for c in ("Fabricante (Óleo)", "Modelo (Óleo)")
        if record.get(c) and c in _conf
    ]
    if _name_confs:
        _conf["Fabricante/Modelo"] = min(_name_confs)

    # --- Viscosidade (kept in its own column, lightly normalized) ---
    record["Viscosidade (Óleo)"] = _fix_visc_spacing(record.get("Viscosidade (Óleo)") or "")

    # The OCR sometimes dumps the brand into the viscosity field (e.g. the person
    # wrote "Cat 15W40" and Azure read it all as Viscosidade = "cat 15440").
    # If the viscosity STARTS with an alphabetic token that is a known oil brand
    # (not a viscosity qualifier like SAE/ISO/VG) AND the rest contains a number,
    # move the brand to Fabricante/Modelo and keep only the grade in Viscosidade.
    _VISC_QUALIFIERS = {"sae", "iso", "vg", "nlgi", "aw", "ep", "hd", "w", "hlp", "clp", "atf"}
    _vm = re.match(r"^([A-Za-zÀ-ú]{2,})\s+(.+)$", record["Viscosidade (Óleo)"])
    if _vm:
        _pfx, _rest = _vm.group(1), _vm.group(2)
        if _norm(_pfx) not in _VISC_QUALIFIERS and re.search(r"\d", _rest):
            _brand, _is_brand = _fuzzy_match_reference(_pfx, _fab_map, cutoff=0.82)
            if _is_brand:
                if not record["Fabricante/Modelo"]:
                    record["Fabricante/Modelo"] = _brand
                    _conf["Fabricante/Modelo"] = min(
                        _conf.get("Viscosidade (Óleo)", 90.0), 85.0
                    )
                record["Viscosidade (Óleo)"] = _rest.strip()

    for c in FORCE_EMPTY_IN_EXCEL:
        record[c] = ""
        _conf.pop(c, None)

    # ── Composite reliability: combine the confidence gathered above with a
    # rule-based plausibility ceiling per cell. The final score is the weakest
    # link (min) — a value that is either low-confidence OR implausible gets
    # flagged. This is what the "Revisar" review list reads.
    for _col in COLETA_COLUMNS:
        if _col in FORCE_EMPTY_IN_EXCEL:
            continue
        _v = record.get(_col, "")
        if not _v:
            continue
        _cap = _field_plausibility_cap(_col, _v)
        _base = _conf.get(_col, 90.0)   # no Azure score recorded → neutral base
        _conf[_col] = min(_base, _cap)

    record["__confidence__"] = _conf
    return record


# ============================================================
# 3) FILTERING: IGNORE NON-FORM PAGES
# ============================================================
def _record_has_signal(rec: dict) -> bool:
    def has_big_number(s: str) -> bool:
        return bool(re.search(r"\b\d{6,}\b", (s or "")))

    if has_big_number(rec.get("Número do Frasco", "")):
        return True
    if has_big_number(rec.get("Código externo", "")):
        return True

    key_fields = [
        "Chassi Série",
        "Tag Frota",
        "Ponto de Coleta / Compartimento",
        "Horímetro/Km/Período",
        "Data da Coleta",
        "Óleo trocado",
        "Viscosidade (Óleo)",
        "Fabricante/Modelo",
        "Horas/Km do Fluído",
    ]
    filled = sum(1 for k in key_fields if (rec.get(k, "") or "").strip())
    return filled >= 3


def _is_probably_form_page(pil_img) -> bool:
    import numpy as np
    from PIL import Image

    img = pil_img.convert("L")
    img = img.resize((600, int(600 * img.height / img.width)), Image.BILINEAR)

    arr  = np.array(img, dtype=np.float32)
    h, w = arr.shape

    left  = arr[:, : w // 2]
    right = arr[:, w // 2 :]

    left_mean  = float(left.mean())
    left_std   = float(left.std())
    right_mean = float(right.mean())

    # left_mean < 225 covers both white-background and yellow-background forms.
    # White blank pages have mean ≈ 250+ with std ≈ 0, so they're still excluded.
    looks_like_form = (left_mean < 225) and (left_std > 20)
    not_blank       = (left_mean < 248) or (right_mean < 248)

    return bool(looks_like_form and not_blank)


# ============================================================
# 4) AZURE DOCUMENT INTELLIGENCE
# ============================================================
ENDPOINT = st.secrets["AZURE_DI_ENDPOINT"]
KEY      = st.secrets["AZURE_DI_KEY"]
MODEL_ID = st.secrets["AZURE_DI_MODEL_ID"]

di_client = DocumentAnalysisClient(ENDPOINT, AzureKeyCredential(KEY))


def analyze_bytes(file_bytes: bytes):
    poller = di_client.begin_analyze_document(MODEL_ID, document=file_bytes)
    return poller.result()


def result_to_records(result, page_bytes: bytes = None, page_mime: str = "application/pdf") -> tuple[list[dict], list[dict]]:
    records: list[dict]         = []
    raw_fields_list: list[dict] = []
    pages = getattr(result, "pages", []) or []
    if getattr(result, "documents", None):
        for doc_idx, doc in enumerate(result.documents):
            fields      = {}
            confidences = {}
            raw         = {}
            for name, field in (doc.fields or {}).items():
                value      = field.value if field.value is not None else field.content
                confidence = getattr(field, "confidence", None)
                fields[name]      = value
                confidences[name] = confidence or 1.0
                mapped_col        = key_to_column(name)
                raw[name] = {
                    "value":      str(value) if value is not None else "",
                    "confidence": round(confidence, 2) if confidence is not None else None,
                    "mapped_to":  mapped_col or "❌ não mapeado",
                }
            page = pages[doc_idx] if doc_idx < len(pages) else None
            rec  = fields_to_record(fields, page=page, confidences=confidences, page_bytes=page_bytes, page_mime=page_mime)
            records.append(rec)
            # Debug: attach page.words summary so the diagnostic can show OCR words
            # near the Horas/Km do Fluído label when the field is blank.
            if page is not None and not rec.get("Horas/Km do Fluído"):
                _words_debug = []
                for _w in (getattr(page, "words", None) or []):
                    _poly = getattr(_w, "polygon", None)
                    if not _poly:
                        continue
                    _ys = [_p.y for _p in _poly]
                    _xs = [_p.x for _p in _poly]
                    _cy = sum(_ys) / len(_ys)
                    _cx = sum(_xs) / len(_xs)
                    _words_debug.append({"t": _w.content, "x": round(_cx, 3), "y": round(_cy, 3)})
                raw["__words_debug__"] = _words_debug
            raw_fields_list.append(raw)
    return records, raw_fields_list


def pdf_split_to_single_page_pdfs(pdf_bytes: bytes) -> list[bytes]:
    import fitz
    src = fitz.open(stream=pdf_bytes, filetype="pdf")
    out: list[bytes] = []
    for i in range(src.page_count):
        dst = fitz.open()
        dst.insert_pdf(src, from_page=i, to_page=i)
        out.append(dst.tobytes())
        dst.close()
    src.close()
    return out


def pdf_single_page_pdf_to_pil(single_page_pdf_bytes: bytes, dpi: int = 72):
    import fitz
    from PIL import Image
    import io
    doc  = fitz.open(stream=single_page_pdf_bytes, filetype="pdf")
    page = doc[0]
    pix  = page.get_pixmap(dpi=dpi, alpha=False)
    doc.close()
    return Image.open(io.BytesIO(pix.tobytes("png")))


def extract_records_from_upload(
    file_bytes: bytes,
    mime_type: str,
    progress_cb=None,
) -> tuple[list[dict], dict, list[dict]]:
    stats = {
        "pdf_pages_total":      0,
        "pdf_pages_candidates": 0,
        "records_before_filter":0,
        "records_after_filter": 0,
        "skipped_pages":        0,
    }

    if mime_type != "application/pdf":
        if progress_cb:
            progress_cb(0.1, "Analisando imagem...")
        result = analyze_bytes(file_bytes)
        if progress_cb:
            progress_cb(1.0, "Concluído")
        records, raw_fields_list = result_to_records(result, page_bytes=file_bytes, page_mime=mime_type)
        stats["records_before_filter"] = len(records)
        filtered       = [(r, rf) for r, rf in zip(records, raw_fields_list) if _record_has_signal(r)]
        records        = [x[0] for x in filtered]
        raw_fields_list= [x[1] for x in filtered]
        stats["records_after_filter"] = len(records)
        return records, stats, raw_fields_list

    page_pdfs = pdf_split_to_single_page_pdfs(file_bytes)
    stats["pdf_pages_total"] = len(page_pdfs)

    candidate_pages: list[bytes] = []
    for i, spdf in enumerate(page_pdfs):
        if progress_cb:
            progress_cb(i / len(page_pdfs) * 0.3, f"Classificando página {i+1}/{len(page_pdfs)}...")
        try:
            pil = pdf_single_page_pdf_to_pil(spdf, dpi=72)
            if _is_probably_form_page(pil):
                candidate_pages.append(spdf)
            else:
                stats["skipped_pages"] += 1
        except Exception:
            candidate_pages.append(spdf)

    stats["pdf_pages_candidates"] = len(candidate_pages)

    all_records: list[dict] = []
    all_raw: list[dict]     = []
    n_cand = len(candidate_pages)
    for i, spdf in enumerate(candidate_pages):
        if progress_cb:
            pct = 0.3 + (i / n_cand) * 0.7
            progress_cb(pct, f"Extraindo formulário {i+1}/{n_cand}...")
        result = analyze_bytes(spdf)
        recs, raws = result_to_records(result, page_bytes=spdf, page_mime="application/pdf")
        all_records.extend(recs)
        all_raw.extend(raws)

    if progress_cb:
        progress_cb(1.0, "Concluído!")

    stats["records_before_filter"] = len(all_records)

    filtered    = [(r, rf) for r, rf in zip(all_records, all_raw) if _record_has_signal(r)]
    cleaned     = [x[0] for x in filtered]
    raw_cleaned = [x[1] for x in filtered]
    stats["records_after_filter"] = len(cleaned)

    return cleaned, stats, raw_cleaned


# ============================================================
# 5) STREAMLIT UI
# ============================================================
st.title("OCR – Cartão de Óleo → Excel")
st.caption("Fluxo: enviar arquivo → extrair → baixar Excel. (Ignora páginas que não são formulário.)")

uploaded_file = st.file_uploader(
    "Envie um cartão (imagem) ou um PDF com várias páginas (formulários misturados)",
    type=["jpg", "jpeg", "png", "pdf"],
)

if uploaded_file is not None and uploaded_file.type.startswith("image/"):
    st.image(uploaded_file, caption="Imagem enviada")

if uploaded_file is None:
    st.info("Envie um arquivo para habilitar a extração.")
    st.stop()

st.markdown("### Extrair e gerar Excel")

if st.button("Extrair somente formulários e baixar Excel"):
    _prog_bar  = st.progress(0, text="Iniciando...")
    _prog_text = st.empty()

    def _on_progress(pct: float, msg: str):
        _prog_bar.progress(min(pct, 1.0), text=msg)
        _prog_text.caption(f"{pct*100:.0f}% — {msg}")

    try:
        records, stats, raw_fields_list = extract_records_from_upload(
            uploaded_file.getvalue(), uploaded_file.type, progress_cb=_on_progress
        )
    except Exception as e:
        st.error(f"Erro ao processar: {e}")
        st.stop()

    _prog_bar.empty()
    _prog_text.empty()

    if uploaded_file.type == "application/pdf":
        st.write(
            f"PDF: {stats['pdf_pages_total']} páginas | "
            f"candidatas (form): {stats['pdf_pages_candidates']} | "
            f"ignoradas: {stats['skipped_pages']}"
        )

    st.write(f"Registros (antes do filtro): {stats['records_before_filter']}")
    st.write(f"Registros (após filtro): **{stats['records_after_filter']}**")

    if not records:
        st.warning("Não consegui extrair registros úteis (ou todas as páginas foram classificadas como não-formulário).")
        st.stop()

    with st.expander("Prévia (primeiras 5 linhas)"):
        for i, rec in enumerate(records[:5], start=1):
            st.markdown(f"**Linha {i}**")
            for c in COLETA_COLUMNS:
                st.write(f"- {c}: {rec.get(c, '')}")
            st.divider()


    import pandas as pd
    # Fonte: score COMPOSTO de confiabilidade por célula (confiança Azure + regra de
    # plausibilidade + catálogo), guardado em record["__confidence__"]. Já em 0-100.
    # Só conta células efetivamente preenchidas (campo vazio não tem confiabilidade).
    _conf_by_col: dict[str, list[float]] = {}
    for _rec in records:
        _rconf = _rec.get("__confidence__", {}) or {}
        for col in COLETA_COLUMNS:
            if col in FORCE_EMPTY_IN_EXCEL:
                continue
            if _rec.get(col) and col in _rconf:
                _conf_by_col.setdefault(col, []).append(_rconf[col])

    _conf_rows = []
    for col in COLETA_COLUMNS:
        confs = _conf_by_col.get(col, [])
        if confs:
            _conf_rows.append({
                "Campo":              col,
                "Confiança média":    round(sum(confs) / len(confs), 1),
                "Amostras":           len(confs),
            })

    if _conf_rows:
        _all_confs = [c for v in _conf_by_col.values() for c in v]
        _conf_rows.append({
            "Campo":           "MÉDIA GERAL",
            "Confiança média": round(sum(_all_confs) / len(_all_confs), 1),
            "Amostras":        len(_all_confs),
        })
        df_conf = pd.DataFrame(_conf_rows)

        def _style_conf(row):
            if row["Campo"] == "MÉDIA GERAL":
                return ["background-color: #BDD7EE; font-weight: bold"] * len(row)
            if row["Confiança média"] < 70:
                return ["background-color: #FCE4D6; color: #C00000"] * len(row)
            if row["Confiança média"] < 85:
                return ["background-color: #FFF2CC"] * len(row)
            return ["background-color: #E2EFDA"] * len(row)

        with st.expander("📊 Confiança média por campo"):
            st.dataframe(
                df_conf.style
                    .format({"Confiança média": "{:.1f}%"})
                    .apply(_style_conf, axis=1),
                use_container_width=True,
                hide_index=True,
            )
            st.caption("Verde ≥ 85% · Amarelo 70–84% · Vermelho < 70%")

    excel_bytes = build_excel_bytes(records)
    st.download_button(
        "Baixar Excel",
        data=excel_bytes,
        file_name="coleta.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
